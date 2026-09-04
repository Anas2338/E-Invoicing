"""
Excel service for template generation, parsing, and validation.
"""
from typing import Optional, Tuple
from datetime import datetime, date, time
import pandas as pd
from io import BytesIO
from pathlib import Path
from uuid import UUID
from sqlmodel import Session, select
import logging

from src.models.excel_upload_session import ExcelUploadSession, ExcelUploadProcessingStatus
from src.schemas.invoice import Environment
from src.utils.excel_validator import ExcelValidator

logger = logging.getLogger(__name__)


def _shade_columns_to_last_row(
    buf: BytesIO,
    all_column_letters: list[str],
    fill_column_letters: list[str],
) -> None:
    """Apply per-column styles for the full column height (rows 1..1048576).

    Excel's ``<col style="...">`` attribute applies the style to every cell in
    that column, which styles the full column height without materializing
    millions of cells. openpyxl does not serialize column styles, so the style
    indices are resolved from styles.xml and injected into the sheet's <col>
    elements after saving.

    Every column gets a thin border; the ``fill_column_letters`` additionally
    get the light blue background. The column styles use plain fonts and no
    alignment so data rows aren't affected by header formatting.
    """
    import re
    import zipfile
    from openpyxl.utils import column_index_from_string

    rgb = '00DDEBF7'
    fill_numbers = {column_index_from_string(letter) for letter in fill_column_letters}
    all_numbers = [column_index_from_string(letter) for letter in all_column_letters]

    with zipfile.ZipFile(BytesIO(buf.getvalue())) as zin:
        styles_xml = zin.read('xl/styles.xml').decode('utf-8')

        # Resolve the fill index in the workbook's style table
        # (fills[0] and fills[1] are the mandatory defaults).
        fills_match = re.search(r'<fills[^>]*>(.*?)</fills>', styles_xml, re.S)
        fill_index = None
        for idx, fill in enumerate(re.finditer(r'<fill>(.*?)</fill>', fills_match.group(1), re.S)):
            if f'rgb="{rgb}"' in fill.group(1):
                fill_index = idx
                break
        if fill_index is None:
            raise RuntimeError('light blue fill not found in workbook styles')

        # Resolve the thin border index (a border with style="thin" sides).
        borders_match = re.search(r'<borders[^>]*>(.*?)</borders>', styles_xml, re.S)
        border_pattern = re.compile(r'<border\b[^>]*?/>|<border\b.*?</border>', re.S)
        thin_border_id = None
        for idx, border_xml in enumerate(border_pattern.findall(borders_match.group(1))):
            if 'style="thin"' in border_xml:
                thin_border_id = idx
                break
        if thin_border_id is None:
            raise RuntimeError('thin border not found in workbook styles')

        # Which fonts are bold? Column styles must not reference them.
        # Match both full <font>…</font> elements and self-closing <font />.
        fonts_match = re.search(r'<fonts[^>]*>(.*?)</fonts>', styles_xml, re.S)
        fonts_bold = []
        if fonts_match:
            font_pattern = re.compile(r'<font\b[^>]*?/>|<font\b.*?</font>', re.S)
            for font_xml in font_pattern.findall(fonts_match.group(1)):
                fonts_bold.append(bool(re.search(r'<b\b', font_xml)))

        # Find plain (non-bold, no alignment) cell styles for the exact
        # fill/border combinations — rows 3+ inherit these via <col>.
        cell_xfs_match = re.search(r'<cellXfs[^>]*>(.*?)</cellXfs>', styles_xml, re.S)
        xfs = [m.group(0) for m in re.finditer(r'<xf\b[^>]*?>', cell_xfs_match.group(1), re.S)]

        def find_style_xf(fill_id: int, border_id: int) -> int | None:
            for xf_idx, xf in enumerate(xfs):
                f_id = int(re.search(r'fillId="(\d+)"', xf).group(1))
                b_id = int(re.search(r'borderId="(\d+)"', xf).group(1))
                font_id = int(re.search(r'fontId="(\d+)"', xf).group(1))
                if (
                    f_id == fill_id
                    and b_id == border_id
                    and not fonts_bold[font_id]
                    and 'applyAlignment' not in xf
                ):
                    return xf_idx
            return None

        fill_border_style = find_style_xf(fill_index, thin_border_id)
        border_only_style = find_style_xf(0, thin_border_id)
        if fill_border_style is None or border_only_style is None:
            raise RuntimeError('plain (fill/border) cell styles not found for shading')

        style_by_column = {
            n: (fill_border_style if n in fill_numbers else border_only_style)
            for n in all_numbers
        }

        def patch_cols(sheet_xml: str) -> str:
            def add_style(match: re.Match) -> str:
                attrs = match.group(1)
                m_num = re.search(r'min="(\d+)"', attrs)
                style = style_by_column[int(m_num.group(1))]
                if 'style=' in attrs:
                    attrs = re.sub(r'style="[^"]*"', f'style="{style}"', attrs)
                else:
                    attrs += f' style="{style}"'
                return f'<col{attrs}/>'

            patched = sheet_xml
            for n in all_numbers:
                patched = re.sub(r'<col\b([^>]*?)/>', add_style, patched)
            return patched

        entries: list[tuple[zipfile.ZipInfo, bytes]] = []
        for item in zin.infolist():
            content = zin.read(item.filename)
            if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                content = patch_cols(content.decode('utf-8')).encode('utf-8')
            entries.append((item, content))

    buf.seek(0)
    buf.truncate()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item, content in entries:
            zout.writestr(item, content)


def _clean_ntn_cnic(value) -> str:
    """Normalize NTN/CNIC value from Excel cell.

    Returns empty string for blank cells and common placeholder values
    that Excel/pandas produce (e.g., 0, 0.0, nan). Handles float-to-int
    conversion for numeric cells (e.g., 1234567.0 → "1234567").
    Otherwise returns the trimmed string representation.
    """
    if pd.isna(value):
        return ""
    # Handle numeric types: convert float to int if it's a whole number,
    # then to string — avoids "1234567.0" from pandas float reading
    if isinstance(value, float):
        if value == int(value):
            value = int(value)
    s = str(value).strip()
    # Treat common empty-like / placeholder values as genuinely empty
    if s in ("", "0", "0.0", "nan", "None", "none", "null", "N/A", "n/a", "-", "NA", "na", "Nil", "nil"):
        return ""
    # Strip trailing ".0" that may come from numeric string conversion
    if s.endswith(".0") and len(s) > 2:
        s = s[:-2]
    return s


def _extract_invoice_number_suffix(value) -> Optional[int]:
    """Extract the trailing numeric part of an invoice number.

    e.g. "INV-0005" -> 5, "INV-2026-0007" -> 7.
    Returns None when there are no trailing digits.
    """
    import re
    if not value:
        return None
    match = re.search(r"(\d+)$", str(value))
    return int(match.group(1)) if match else None


def _format_invoice_number(
    prefix: str,
    number: int,
    padding: int = 4,
    include_year: bool = False,
) -> str:
    """Format a numeric sequence into the user's configured invoice number format.

    e.g. _format_invoice_number("INV-", 6, 4) -> "INV-0006"
         _format_invoice_number("INV-", 6, 4, True) -> "INV-2026-0006"
    """
    padded = str(number).zfill(padding)
    if include_year:
        return f"{prefix}{datetime.now().year}-{padded}"
    return f"{prefix}{padded}"


class _AutoInvoiceNumberGenerator:
    """Sequential invoice number generator for rows with a blank invoice_number cell.

    Mirrors the main backend's logic (backend/src/utils/helpers.py
    get_next_invoice_number and backend/src/utils/manual_excel_helper.py
    _generate_auto_invoice_numbers): the sequence starts at the configured
    start number (or one after the highest trailing-digit suffix among the
    taken numbers) and skips any number already in use. `taken` contains
    existing invoice numbers from both the automation DB and the main DB,
    plus explicit numbers present in the uploaded file.
    """

    def __init__(
        self,
        prefix: str = "INV-",
        start_number: int = 1,
        padding: int = 4,
        include_year: bool = False,
        taken: Optional[set[str]] = None,
    ):
        self.prefix = prefix or "INV-"
        self.padding = padding or 4
        self.include_year = include_year
        self.taken = set(taken or [])

        max_suffix = (start_number or 1) - 1
        for candidate in self.taken:
            suffix = _extract_invoice_number_suffix(candidate)
            if suffix is not None and suffix > max_suffix:
                max_suffix = suffix
        self.next_number = max_suffix + 1

    def next(self) -> str:
        """Return the next available invoice number and mark it as taken."""
        candidate = _format_invoice_number(self.prefix, self.next_number, self.padding, self.include_year)
        while candidate in self.taken:
            self.next_number += 1
            candidate = _format_invoice_number(self.prefix, self.next_number, self.padding, self.include_year)
        self.taken.add(candidate)
        self.next_number += 1
        return candidate


class ExcelService:
    """Service for Excel file operations."""

    TEMPLATE_COLUMNS = [
        # Invoice identification (invoice_number is intentionally absent:
        # numbers are assigned by the transfer job at schedule time, not on upload)
        "invoice_type",
        "invoice_date",

        # Buyer information
        "buyer_ntn_cnic",
        "buyer_business_name",
        "buyer_province",
        "buyer_address",
        "buyer_registration_type",

        # Item details - simplified with saved_item_code
        "saved_item_code",
        "product_description",
        "quantity",
        "value_sales_excluding_st",
        "fixed_notified_value_or_retail_price",
        "further_tax",
        "discount",

        # Income tax
        "income_tax",
        "withholding_tax_amount",

        # Scheduling
        "scheduled_date",
        "scheduled_time",

    ]

    # Manual invoice Excel template columns
    # Based on automation template but without scheduled_date/scheduled_time, with income_tax
    MANUAL_TEMPLATE_COLUMNS = [
        # Invoice identification
        "invoice_number",
        "invoice_type",
        "invoice_date",

        # Buyer information
        "buyer_ntn_cnic",
        "buyer_business_name",
        "buyer_province",
        "buyer_address",
        "buyer_registration_type",

        # Item details - simplified with saved_item_code
        "saved_item_code",
        "quantity",
        "value_sales_excluding_st",
        "fixed_notified_value_or_retail_price",
        "further_tax",
        "discount",

        # Income tax
        "income_tax",
        "withholding_tax_amount",
    ]

    def __init__(self, db: Session):
        """
        Initialize Excel service.

        Args:
            db: Database session
        """
        self.db = db
        self.validator = ExcelValidator()

    def generate_excel_template(self) -> BytesIO:
        """
        Generate Excel template with predefined headers.
        Simplified template using saved_item_code for auto-population.

        Returns:
            BytesIO object containing Excel file
        """
        # Create DataFrame with column headers
        df = pd.DataFrame(columns=self.TEMPLATE_COLUMNS)

        # Add a demo sample row so users can see the expected format.
        # The parser skips this row on upload by matching the demo markers
        # (buyer "ABC Corporation" + saved_item_code "ITEM001").
        sample_row = pd.DataFrame([{
            "invoice_type": "Sale Invoice",
            "invoice_date": "2026-05-12",
            "buyer_ntn_cnic": "1234567",
            "buyer_business_name": "ABC Corporation",
            "buyer_province": "PUNJAB",
            "buyer_address": "123 Main Street, Lahore",
            "buyer_registration_type": "Registered",
            "saved_item_code": "ITEM001",
            "product_description": "Widget - Model X",
            "quantity": "2",
            "value_sales_excluding_st": "50000",
            "fixed_notified_value_or_retail_price": "0",
            "further_tax": "0",
            "discount": "0",
            "income_tax": "236G",
            "withholding_tax_amount": "50",
            "scheduled_date": "2026-05-13",
            "scheduled_time": "10:00",
        }])
        df = pd.concat([df, sample_row], ignore_index=True)

        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Invoices')

            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Invoices']

            # Set column widths for better readability
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.datavalidation import DataValidation

            column_widths = [
                15,  # invoice_type
                12,  # invoice_date
                15,  # buyer_ntn_cnic
                25,  # buyer_business_name
                15,  # buyer_province
                30,  # buyer_address
                20,  # buyer_registration_type
                20,  # saved_item_code
                20,  # product_description
                10,  # quantity
                20,  # value_sales_excluding_st
                25,  # fixed_notified_value_or_retail_price
                12,  # further_tax
                12,  # discount
                12,  # income_tax
                20,  # withholding_tax_amount
                15,  # scheduled_date
                15,  # scheduled_time
            ]

            for idx, width in enumerate(column_widths, start=1):
                col_letter = get_column_letter(idx)
                worksheet.column_dimensions[col_letter].width = width

            # Freeze the header row
            worksheet.freeze_panes = 'A2'

            # — Light blue fill + thin borders across the full sheet —
            from openpyxl.styles import Font, PatternFill, Border, Side

            # Bold headings only in the header row — data rows stay plain via
            # the column styles patched below.
            bold_font = Font(bold=True)
            thin_side = Side(style='thin')
            thin_border = Border(
                left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
            )
            for col_idx in range(1, len(column_widths) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = bold_font
                cell.border = thin_border
            # Register the thin-border style on the demo row too, so the column
            # patch below can reference it for rows 3+ (full column height).
            for col_idx in range(1, len(column_widths) + 1):
                worksheet.cell(row=2, column=col_idx).border = thin_border

            light_blue_fill = PatternFill(
                start_color='DDEBF7', end_color='DDEBF7', fill_type='solid'
            )
            highlighted_columns = {
                "invoice_type", "invoice_date", "buyer_business_name",
                "buyer_province", "buyer_address", "buyer_registration_type",
                "saved_item_code", "product_description", "quantity",
                "value_sales_excluding_st", "scheduled_date", "scheduled_time",
            }
            highlighted_letters = sorted(
                get_column_letter(self.TEMPLATE_COLUMNS.index(name) + 1)
                for name in highlighted_columns
            )
            for name in highlighted_columns:
                letter = get_column_letter(self.TEMPLATE_COLUMNS.index(name) + 1)
                # Fill the header cell so openpyxl registers the style, then
                # patch the <col> elements below for the full column height.
                worksheet[f'{letter}1'].fill = light_blue_fill
                # Fill the demo row cell explicitly so the sample line shows
                # the colour in every viewer (rows 3+ rely on <col> style).
                worksheet[f'{letter}2'].fill = light_blue_fill

            all_column_letters = [
                get_column_letter(i) for i in range(1, len(self.TEMPLATE_COLUMNS) + 1)
            ]

            # Data validation dropdowns
            last_data_row = 1048576  # covers entire column

            option_sets: dict[str, list[str]] = {
                'A': ["Sale Invoice", "Debit Note", "Credit Note"],        # invoice_type
                'E': ["PUNJAB", "SINDH", "KPK", "BALOCHISTAN", "ISLAMABAD", "GILGIT BALTISTAN", "AZAD JAMMU KASHMIR"],  # buyer_province
                'G': ["Registered", "Unregistered", "Final Consumer"],     # buyer_registration_type
                'O': ["236G", "236H", "None"],                              # income_tax
            }

            for col_letter, values in option_sets.items():
                options_str = ','.join(values)
                dv = DataValidation(
                    type='list',
                    formula1=f'"{options_str}"',
                    allow_blank=True,
                )
                dv.error = 'Please select a valid value from the dropdown list.'
                dv.errorTitle = 'Invalid value'
                dv.prompt = 'Select from the dropdown list'
                dv.promptTitle = 'Valid options'
                worksheet.add_data_validation(dv)
                dv.add(f'{col_letter}2:{col_letter}{last_data_row}')

        _shade_columns_to_last_row(output, all_column_letters, highlighted_letters)

        output.seek(0)
        return output

    def generate_manual_excel_template(self) -> BytesIO:
        """
        Generate Excel template for manual invoice upload.
        Based on automation template but without scheduled_date/scheduled_time,
        and with income_tax column (236G or 236H).

        Returns:
            BytesIO object containing Excel file
        """
        df = pd.DataFrame(columns=self.MANUAL_TEMPLATE_COLUMNS)

        sample_row = pd.DataFrame([{
            "invoice_number": "INV-001",
            "invoice_type": "Sale Invoice",
            "invoice_date": "2026-05-12",
            "buyer_ntn_cnic": "1234567",
            "buyer_business_name": "ABC Corporation",
            "buyer_province": "PUNJAB",
            "buyer_address": "123 Main Street, Lahore",
            "buyer_registration_type": "Registered",
            "saved_item_code": "ITEM001",
            "quantity": "2",
            "value_sales_excluding_st": "50000",
            "fixed_notified_value_or_retail_price": "0",
            "further_tax": "0",
            "discount": "0",
            "income_tax": "236G",
            "withholding_tax_amount": "50",
        }])
        df = pd.concat([df, sample_row], ignore_index=True)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Invoices')

            workbook = writer.book
            worksheet = writer.sheets['Invoices']

            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.datavalidation import DataValidation

            column_widths = [
                15,  # invoice_number
                15,  # invoice_type
                12,  # invoice_date
                15,  # buyer_ntn_cnic
                25,  # buyer_business_name
                15,  # buyer_province
                30,  # buyer_address
                20,  # buyer_registration_type
                20,  # saved_item_code
                10,  # quantity
                20,  # value_sales_excluding_st
                25,  # fixed_notified_value_or_retail_price
                12,  # further_tax
                12,  # discount
                12,  # income_tax
                20,  # withholding_tax_amount
            ]

            for idx, width in enumerate(column_widths, start=1):
                col_letter = get_column_letter(idx)
                worksheet.column_dimensions[col_letter].width = width

            # Freeze the header row
            worksheet.freeze_panes = 'A2'

            # Bold formatting for header row
            from openpyxl.styles import Font
            bold_font = Font(bold=True)
            for col_idx in range(1, len(column_widths) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = bold_font

            # Data validation dropdowns
            last_data_row = 1048576

            option_sets: dict[str, list[str]] = {
                'B': ["Sale Invoice", "Debit Note", "Credit Note"],
                'F': ["PUNJAB", "SINDH", "KPK", "BALOCHISTAN", "ISLAMABAD", "GILGIT BALTISTAN", "AZAD JAMMU KASHMIR"],
                'H': ["Registered", "Unregistered", "Final Consumer"],
                'O': ["236G", "236H", "None"],
            }

            for col_letter, values in option_sets.items():
                options_str = ','.join(values)
                dv = DataValidation(
                    type='list',
                    formula1=f'"{options_str}"',
                    allow_blank=True,
                )
                dv.error = 'Please select a valid value from the dropdown list.'
                dv.errorTitle = 'Invalid value'
                dv.prompt = 'Select from the dropdown list'
                dv.promptTitle = 'Valid options'
                worksheet.add_data_validation(dv)
                dv.add(f'{col_letter}2:{col_letter}{last_data_row}')

        output.seek(0)
        return output

    def validate_excel_structure(self, file_source: BytesIO | str) -> Tuple[bool, Optional[str]]:
        """
        Validate Excel file structure.

        Args:
            file_source: Path to Excel file or BytesIO object

        Returns:
            Tuple of (is_valid, error_message)
        """
        return self.validator.validate_excel_structure(file_source)

    def check_duplicate_invoices(self, file_source: BytesIO | str) -> Tuple[bool, Optional[str]]:
        """
        Check for duplicate invoice numbers in Excel file.

        Args:
            file_source: Path to Excel file or BytesIO object

        Returns:
            Tuple of (is_valid, error_message)
        """
        return self.validator.validate_no_duplicate_invoices(file_source)

    def parse_excel_file(self, file_source: BytesIO | str, user_id: UUID = None, main_db: Session = None) -> list[dict]:
        """
        Parse Excel file and return invoice data.
        Auto-populates item details from saved_item_code and seller info from user's business information.

        The template has no invoice_number column — invoice numbers are assigned
        by the transfer job when each invoice's scheduled time arrives
        (see InvoiceNumberAssigner in transfer_service.py).

        Args:
            file_source: Path to Excel file or BytesIO object
            user_id: User UUID for fetching saved items and seller info
            main_db: Main database session for fetching user information

        Returns:
            List of invoice dictionaries with FBR-compliant structure

        Raises:
            ValueError: If Excel file is invalid or saved_item_code not found
            MemoryError: If file is too large to process in memory
        """
        try:
            # Read Excel file (pandas handles both str and BytesIO)
            try:
                df = pd.read_excel(
                    file_source,
                    engine='openpyxl',
                    dtype={
                        'saved_item_code': str
                    }
                )
            except MemoryError:
                raise MemoryError(
                    "File is too large to process in memory. "
                    "Please reduce the file size or split into smaller batches (max 1,000 rows)."
                )

            # Remove fully-empty rows
            df = df.dropna(how='all')

            # Skip the demo sample row (buyer "ABC Corporation" + saved_item_code
            # "ITEM001") if the user left it in the template
            if not df.empty:
                demo_mask = (
                    df['buyer_business_name'].astype(str).str.strip() == 'ABC Corporation'
                ) & (df['saved_item_code'].astype(str).str.strip() == 'ITEM001')
                df = df[~demo_mask]

            # Fetch user's seller information from main database
            from src.models.user import User
            seller_info = {}
            user = None
            if user_id and main_db:
                user = main_db.get(User, user_id)
                if user:
                    seller_info = {
                        "seller_ntn_cnic": user.fbr_seller_ntn or "",
                        "seller_business_name": user.fbr_business_name or "",
                        "seller_province": user.fbr_seller_province or "",
                        "seller_address": user.fbr_seller_address or "",
                    }

            # Fetch all user's saved items for lookup from main database
            from src.models.user_saved_product import UserSavedProduct
            saved_items_dict = {}
            if user_id and main_db:
                from sqlmodel import select
                statement = select(UserSavedProduct).where(
                    UserSavedProduct.user_id == user_id,
                    UserSavedProduct.is_active == 1
                )
                saved_items = main_db.exec(statement).all()
                saved_items_dict = {item.item_code: item for item in saved_items}

            # Fetch UOM descriptions from FBR master data
            from src.models.fbr_master_data import FBRUOM, FBRTransactionType
            uom_descriptions = {}
            transaction_type_descriptions = {}
            if main_db:
                try:
                    uom_records = main_db.query(FBRUOM).all()
                    uom_descriptions = {uom.code: uom.name for uom in uom_records}
                except Exception as e:
                    logger.warning(f"Failed to fetch UOM descriptions: {str(e)}")
                    uom_descriptions = {}

                try:
                    transaction_type_records = main_db.query(FBRTransactionType).all()
                    transaction_type_descriptions = {tt.code: tt.name for tt in transaction_type_records}
                except Exception as e:
                    logger.warning(f"Failed to fetch transaction type descriptions: {str(e)}")
                    transaction_type_descriptions = {}

            # Note: no invoice number handling here — the automation template no
            # longer has an invoice_number column. Numbers are assigned by the
            # transfer job at schedule time (see InvoiceNumberAssigner).

            # Convert to list of dictionaries
            invoices = []
            validation_errors = []  # Collect all validation errors
            for row_idx, row in df.iterrows():
                # Get saved_item_code
                saved_item_code = str(row['saved_item_code']).strip() if pd.notna(row['saved_item_code']) else ""

                # Fetch saved item details
                if not saved_item_code:
                    validation_errors.append(f"Row {row_idx + 2}: saved_item_code is required")
                    continue

                saved_item = saved_items_dict.get(saved_item_code)
                if not saved_item:
                    validation_errors.append(f"Row {row_idx + 2}: saved_item_code '{saved_item_code}' not found in your saved items")
                    continue

                # Validate invoice_type
                VALID_INVOICE_TYPES = ["Sale Invoice", "Debit Note", "Credit Note"]
                invoice_type_raw = str(row['invoice_type']).strip() if pd.notna(row['invoice_type']) else ""
                if invoice_type_raw and invoice_type_raw not in VALID_INVOICE_TYPES:
                    validation_errors.append(
                        f"Row {row_idx + 2}: "
                        f"invoice_type '{invoice_type_raw}' is invalid. Must be one of: {', '.join(VALID_INVOICE_TYPES)}."
                    )
                    continue

                # Validate buyer_province
                VALID_PROVINCES = ["PUNJAB", "SINDH", "KPK", "BALOCHISTAN", "ISLAMABAD", "GILGIT BALTISTAN", "AZAD JAMMU KASHMIR"]
                buyer_province_raw = str(row['buyer_province']).strip() if pd.notna(row['buyer_province']) else ""
                if buyer_province_raw and buyer_province_raw not in VALID_PROVINCES:
                    validation_errors.append(
                        f"Row {row_idx + 2}: "
                        f"buyer_province '{buyer_province_raw}' is invalid. Must be one of: {', '.join(VALID_PROVINCES)}."
                    )
                    continue

                # Validate buyer_registration_type
                VALID_REG_TYPES = ["Registered", "Unregistered", "Final Consumer"]
                buyer_reg_type_raw = str(row['buyer_registration_type']).strip() if pd.notna(row['buyer_registration_type']) else ""
                if buyer_reg_type_raw and buyer_reg_type_raw not in VALID_REG_TYPES:
                    validation_errors.append(
                        f"Row {row_idx + 2}: "
                        f"buyer_registration_type '{buyer_reg_type_raw}' is invalid. Must be one of: {', '.join(VALID_REG_TYPES)}."
                    )
                    continue

                # Parse numeric fields from Excel (rounded to 2 decimal places)
                quantity = round(float(row['quantity']) if pd.notna(row['quantity']) else 0, 2)
                value_sales_excluding_st = round(float(row['value_sales_excluding_st']) if pd.notna(row['value_sales_excluding_st']) else 0, 2)
                fixed_notified_value_or_retail_price = round(float(row['fixed_notified_value_or_retail_price']) if pd.notna(row['fixed_notified_value_or_retail_price']) else 0, 2)
                further_tax = round(float(row['further_tax']) if pd.notna(row['further_tax']) else 0, 2)
                discount = round(float(row['discount']) if pd.notna(row.get('discount')) else 0, 2)

                # Parse income_tax (must be "236G", "236H", or "None")
                income_tax = "236G"
                if pd.notna(row.get('income_tax')):
                    income_tax_raw = str(row['income_tax']).strip()
                    if income_tax_raw in ("236G", "236H", "None"):
                        income_tax = income_tax_raw
                    elif income_tax_raw:
                        validation_errors.append(
                            f"Row {row_idx + 2}: "
                            f"income_tax '{income_tax_raw}' is invalid. Must be '236G', '236H', or 'None'."
                        )
                        continue

                # Parse withholding_tax_amount from Excel (optional, auto-calc if omitted)
                withholding_tax_amount = None
                if pd.notna(row.get('withholding_tax_amount')):
                    try:
                        withholding_tax_amount = round(float(row['withholding_tax_amount']), 2)
                    except (ValueError, TypeError):
                        withholding_tax_amount = None
                if withholding_tax_amount is None:
                    if income_tax == "236H":
                        wht_rate = 0.005
                    elif income_tax == "236G":
                        wht_rate = 0.001
                    else:
                        wht_rate = 0  # "None" - no withholding tax
                    withholding_tax_amount = round(value_sales_excluding_st * wht_rate, 2)

                # Calculate sales tax based on saved item's tax rate
                tax_rate = float(saved_item.default_rate) if saved_item.default_rate else 18.0
                # Use the greater value between Value Excl. Tax and Fixed/Retail Price (FBR rule)
                base_value = max(value_sales_excluding_st, fixed_notified_value_or_retail_price)
                sales_tax_applicable = round((base_value * tax_rate) / 100, 2)
                total_values = round(base_value + sales_tax_applicable + further_tax - discount, 2)

                # Get UOM code and description
                uom_code = saved_item.default_uom or "NOS"
                uom_description = uom_descriptions.get(uom_code, uom_code)

                # Get transaction type (sale type) code and description
                transaction_type_code = saved_item.transaction_type or "01"
                sale_type_description = transaction_type_descriptions.get(transaction_type_code, transaction_type_code)

                # Parse invoice_date to YYYY-MM-DD format
                invoice_date_str = ""
                if pd.notna(row['invoice_date']):
                    try:
                        invoice_date_parsed = pd.to_datetime(row['invoice_date'])
                        invoice_date_str = invoice_date_parsed.strftime('%Y-%m-%d')
                    except:
                        invoice_date_str = str(row['invoice_date']).strip()

                # Product description: use the Excel cell if provided, otherwise
                # fall back to the saved item's description
                product_description = ""
                if pd.notna(row.get('product_description')):
                    product_description = str(row['product_description']).strip()
                if not product_description:
                    product_description = saved_item.product_description or ""

                # Build FBR-compliant invoice data structure.
                # No invoice_number key — assigned by the transfer job at schedule time.
                invoice_data = {
                    "invoice_type": str(row['invoice_type']).strip() if pd.notna(row['invoice_type']) else "Sale Invoice",
                    "invoice_date": invoice_date_str,

                    # Seller information - auto-populated from user's business info
                    "seller_ntn_cnic": seller_info.get("seller_ntn_cnic", ""),
                    "seller_business_name": seller_info.get("seller_business_name", ""),
                    "seller_province": seller_info.get("seller_province", ""),
                    "seller_address": seller_info.get("seller_address", ""),

                    # Buyer information
                    "buyer_ntn_cnic": _clean_ntn_cnic(row['buyer_ntn_cnic']),
                    "buyer_business_name": str(row['buyer_business_name']).strip() if pd.notna(row['buyer_business_name']) else "",
                    "buyer_province": str(row['buyer_province']).strip() if pd.notna(row['buyer_province']) else "",
                    "buyer_address": str(row['buyer_address']).strip() if pd.notna(row['buyer_address']) else "",
                    "buyer_registration_type": str(row['buyer_registration_type']).strip() if pd.notna(row['buyer_registration_type']) else "Registered",

                    # Item details - auto-populated from saved item
                    "items": [{
                        "hs_code": saved_item.hs_code,
                        "product_description": product_description,
                        "rate": saved_item.default_rate or "18",
                        "tax_rate": saved_item.default_rate or "18",
                        "uom": uom_code,
                        "uom_description": uom_description,  # Add description for frontend display
                        "quantity": quantity,
                        "total_values": total_values,
                        "value_sales_excluding_st": value_sales_excluding_st,
                        "fixed_notified_value_or_retail_price": fixed_notified_value_or_retail_price,
                        "sales_tax_applicable": sales_tax_applicable,
                        "sales_tax_withheld_at_source": 0,
                        "extra_tax": 0,
                        "further_tax": further_tax,
                        "sro_schedule_no": saved_item.sro_schedule_no or "",
                        "fed_payable": 0,
                        "discount": discount,
                        "sale_type": saved_item.transaction_type or "01",
                        "sale_type_description": sale_type_description,  # Add description for frontend display
                        "sro_item_serial_no": saved_item.sro_item_serial_no or "",
                        "transaction_type": saved_item.transaction_type or "",
                        # Internal fields (not sent to FBR)
                        "income_tax_type": income_tax,
                        "withholding_tax_amount": withholding_tax_amount,
                    }],

                    # Optional fields
                    "invoice_ref_no": "",
                    "scenario_id": "",

                    # Income tax
                    "income_tax": income_tax,

                    # Environment - always use PRODUCTION for automation
                    "environment": Environment.PRODUCTION,
                }

                # SECURITY: Validate invoice data before adding to list
                from src.utils.invoice_validator import InvoiceValidator
                is_valid, validation_error = InvoiceValidator.validate_invoice_data(invoice_data)

                if not is_valid:
                    # Excel row = pandas index + 2, accounting for the header row
                    logger.warning(f"Invoice validation failed for row {row_idx + 2}: {validation_error}")
                    validation_errors.append(f"Row {row_idx + 2}: {validation_error}")
                    continue

                # Parse scheduling information
                scheduled_date = pd.to_datetime(row['scheduled_date']).date() if pd.notna(row['scheduled_date']) else None
                scheduled_time_str = str(row['scheduled_time']).strip() if pd.notna(row['scheduled_time']) else "00:00"

                # Handle time parsing (could be "10:00" or datetime)
                try:
                    if isinstance(row['scheduled_time'], time):
                        scheduled_time = row['scheduled_time']
                    else:
                        scheduled_time = pd.to_datetime(scheduled_time_str, format='%H:%M').time()
                except:
                    scheduled_time = time(0, 0)

                invoices.append({
                    "invoice_data": invoice_data,
                    "scheduled_date": scheduled_date,
                    "scheduled_time": scheduled_time,
                })

            # If we have validation errors, report them to the user
            if validation_errors:
                error_summary = f"Found {len(validation_errors)} validation error(s):\n" + "\n".join(validation_errors[:5])
                if len(validation_errors) > 5:
                    error_summary += f"\n... and {len(validation_errors) - 5} more errors"
                raise ValueError(error_summary)

            return invoices

        except Exception as e:
            raise ValueError(f"Error parsing Excel file: {str(e)}")

    def parse_excel_for_manual_invoice(self, file_source: BytesIO | str, user_id: UUID = None, main_db: Session = None) -> list[dict]:
        """
        Parse Excel file for manual invoice creation.
        Auto-populates item details from saved_item_code and seller info from user's business information.
        Validates invoice_date is today or previous date (no future dates allowed).
        Does NOT include scheduling fields (those are automation-only).

        Args:
            file_source: Path to Excel file or BytesIO object
            user_id: User UUID for fetching saved items and seller info
            main_db: Main database session for fetching user information

        Returns:
            List of InvoiceCreate-compatible dictionaries

        Raises:
            ValueError: If Excel file is invalid, saved_item_code not found, or date is in the future
        """
        from datetime import date as date_type

        try:
            df = pd.read_excel(
                file_source,
                engine='openpyxl',
                dtype={
                    'saved_item_code': str,
                    'income_tax': str
                }
            )
        except MemoryError:
            raise MemoryError(
                "File is too large to process in memory. "
                "Please reduce the file size or split into smaller batches (max 1,000 rows)."
            )

        df = df.dropna(subset=['invoice_number'])

        if not df.empty:
            df = df[df['invoice_number'].astype(str).str.strip() != 'INV-001']

        # Fetch seller info and saved items (same as automation)
        from src.models.user import User
        seller_info = {}
        if user_id and main_db:
            user = main_db.get(User, user_id)
            if user:
                seller_info = {
                    "seller_ntn_cnic": user.fbr_seller_ntn or "",
                    "seller_business_name": user.fbr_business_name or "",
                    "seller_province": user.fbr_seller_province or "",
                    "seller_address": user.fbr_seller_address or "",
                }

        from src.models.user_saved_product import UserSavedProduct
        saved_items_dict = {}
        if user_id and main_db:
            statement = select(UserSavedProduct).where(
                UserSavedProduct.user_id == user_id,
                UserSavedProduct.is_active == 1
            )
            saved_items = main_db.exec(statement).all()
            saved_items_dict = {item.item_code: item for item in saved_items}

        # Collect all unique invoice numbers from the Excel file
        excel_invoice_numbers = set()
        for _, row in df.iterrows():
            inv_num = str(row['invoice_number']).strip() if pd.notna(row['invoice_number']) else ""
            if inv_num and inv_num != 'INV-001':
                excel_invoice_numbers.add(inv_num)

        # Check which invoice numbers already exist in the database
        existing_invoice_numbers = set()
        if excel_invoice_numbers and main_db:
            from src.models.invoice import Invoice as InvoiceModel
            existing = main_db.exec(
                select(InvoiceModel.external_id).where(
                    InvoiceModel.external_id.in_(excel_invoice_numbers),
                    InvoiceModel.is_deleted == False
                )
            ).all()
            existing_invoice_numbers = set(existing)

        today = date_type.today()
        # Use dict to group rows by invoice_number
        invoice_groups: dict[str, dict] = {}
        validation_errors = []

        for row_idx, row in df.iterrows():
            invoice_number = str(row['invoice_number']).strip()
            excel_row = row_idx + 2  # Human-readable row number (header + 1-indexed)

            # Reject if invoice number already exists in database
            if invoice_number in existing_invoice_numbers:
                validation_errors.append(
                    f"Row {excel_row} (Invoice {invoice_number}): "
                    f"invoice number already exists in your history. Please use a unique invoice number."
                )
                continue

            saved_item_code = str(row['saved_item_code']).strip() if pd.notna(row['saved_item_code']) else ""

            if not saved_item_code:
                validation_errors.append(f"Row {excel_row} (Invoice {invoice_number}): saved_item_code is required")
                continue

            saved_item = saved_items_dict.get(saved_item_code)
            if not saved_item:
                validation_errors.append(f"Row {excel_row} (Invoice {invoice_number}): saved_item_code '{saved_item_code}' not found in your saved items")
                continue

            # Parse and validate invoice_date (must be today or previous, no future)
            invoice_date_str = ""
            if pd.notna(row['invoice_date']):
                try:
                    invoice_date_parsed = pd.to_datetime(row['invoice_date'])
                    invoice_date_str = invoice_date_parsed.strftime('%Y-%m-%d')
                    invoice_date_obj = invoice_date_parsed.date()
                    if invoice_date_obj > today:
                        validation_errors.append(
                            f"Row {excel_row} (Invoice {invoice_number}): "
                            f"invoice_date '{invoice_date_str}' is in the future. Only today or previous dates are allowed."
                        )
                        continue
                except Exception:
                    invoice_date_str = str(row['invoice_date']).strip()
            else:
                validation_errors.append(f"Row {excel_row} (Invoice {invoice_number}): invoice_date is required")
                continue

            # Parse income_tax (must be "236G" or "236H") — first row's value wins for the group
            income_tax = "236G"
            if pd.notna(row.get('income_tax')):
                income_tax_raw = str(row['income_tax']).strip()
                if income_tax_raw in ("236G", "236H"):
                    income_tax = income_tax_raw
                elif income_tax_raw:
                    validation_errors.append(
                        f"Row {excel_row} (Invoice {invoice_number}): "
                        f"income_tax '{income_tax_raw}' is invalid. Must be '236G' or '236H'."
                    )
                    continue

            # Parse withholding_tax_amount from Excel (optional, auto-calc if omitted)
            withholding_tax_amount = None
            if pd.notna(row.get('withholding_tax_amount')):
                try:
                    withholding_tax_amount = round(float(row['withholding_tax_amount']), 2)
                except (ValueError, TypeError):
                    withholding_tax_amount = None

            # Parse numeric fields (rounded to 2 decimal places)
            quantity = round(float(row['quantity']) if pd.notna(row['quantity']) else 0, 2)
            value_sales_excluding_st = round(float(row['value_sales_excluding_st']) if pd.notna(row['value_sales_excluding_st']) else 0, 2)
            fixed_notified_value_or_retail_price = round(float(row['fixed_notified_value_or_retail_price']) if pd.notna(row['fixed_notified_value_or_retail_price']) else 0, 2)
            further_tax = round(float(row['further_tax']) if pd.notna(row['further_tax']) else 0, 2)
            discount = round(float(row['discount']) if pd.notna(row.get('discount')) else 0, 2)

            # Auto-calc WHT if not provided
            if withholding_tax_amount is None:
                if income_tax == "236H":
                    wht_rate = 0.005
                elif income_tax == "236G":
                    wht_rate = 0.001
                else:
                    wht_rate = 0  # "None" - no withholding tax
                withholding_tax_amount = round(value_sales_excluding_st * wht_rate, 2)

            tax_rate = float(saved_item.default_rate) if saved_item.default_rate else 18.0
            # Use the greater value between Value Excl. Tax and Fixed/Retail Price (FBR rule)
            base_value = max(value_sales_excluding_st, fixed_notified_value_or_retail_price)
            sales_tax_applicable = round((base_value * tax_rate) / 100, 2)
            total_values = round(base_value + sales_tax_applicable + further_tax - discount, 2)

            uom_code = saved_item.default_uom or "NOS"

            # Build item dict
            item = {
                "hs_code": saved_item.hs_code,
                "product_description": saved_item.product_description,
                "rate": str(saved_item.default_rate or "18"),
                "uom": uom_code,
                "quantity": quantity,
                "total_values": total_values,
                "value_sales_excluding_st": value_sales_excluding_st,
                "fixed_notified_value_or_retail_price": fixed_notified_value_or_retail_price,
                "sales_tax_applicable": sales_tax_applicable,
                "sales_tax_withheld_at_source": 0,
                "extra_tax": 0,
                "further_tax": further_tax,
                "sro_schedule_no": saved_item.sro_schedule_no or "",
                "fed_payable": 0,
                "discount": discount,
                "sale_type": saved_item.transaction_type or "01",
                "sro_item_serial_no": saved_item.sro_item_serial_no or "",
                # Internal fields (not sent to FBR)
                "income_tax_type": income_tax,
                "withholding_tax_amount": withholding_tax_amount,
            }

            if invoice_number not in invoice_groups:
                # First row for this invoice — capture header-level fields
                invoice_groups[invoice_number] = {
                    "external_id": invoice_number,
                    "invoice_type": str(row['invoice_type']).strip() if pd.notna(row['invoice_type']) else "Sale Invoice",
                    "invoice_date": invoice_date_str,
                    "transaction_type_id": saved_item.transaction_type or "01",
                    "seller_ntn_cnic": seller_info.get("seller_ntn_cnic", ""),
                    "seller_business_name": seller_info.get("seller_business_name", ""),
                    "seller_province": seller_info.get("seller_province", ""),
                    "seller_address": seller_info.get("seller_address", ""),
                    "buyer_ntn_cnic": _clean_ntn_cnic(row['buyer_ntn_cnic']),
                    "buyer_business_name": str(row['buyer_business_name']).strip() if pd.notna(row['buyer_business_name']) else "",
                    "buyer_province": str(row['buyer_province']).strip() if pd.notna(row['buyer_province']) else "",
                    "buyer_address": str(row['buyer_address']).strip() if pd.notna(row['buyer_address']) else "",
                    "buyer_registration_type": str(row['buyer_registration_type']).strip() if pd.notna(row['buyer_registration_type']) else "Registered",
                    "invoice_ref_no": "",
                    "scenario_id": "",
                    "items": [],
                    "environment": Environment.PRODUCTION,
                    "income_tax": income_tax,
                }
            else:
                # Subsequent row — just append the item
                # Validate invoice_date consistency across rows
                existing_date = invoice_groups[invoice_number]["invoice_date"]
                if invoice_date_str != existing_date:
                    validation_errors.append(
                        f"Row {excel_row} (Invoice {invoice_number}): "
                        f"invoice_date '{invoice_date_str}' does not match row's date '{existing_date}'."
                    )
                    continue

            invoice_groups[invoice_number]["items"].append(item)

        if validation_errors:
            error_summary = f"Found {len(validation_errors)} validation error(s):\n" + "\n".join(validation_errors[:5])
            if len(validation_errors) > 5:
                error_summary += f"\n... and {len(validation_errors) - 5} more errors"
            raise ValueError(error_summary)

        # Convert grouped dicts to list, preserving Excel row order (first occurrence)
        invoices = list(invoice_groups.values())
        return invoices

    def check_concurrent_upload(self, user_id: UUID) -> Optional[ExcelUploadSession]:
        """
        Check if user has a concurrent upload in progress.

        Args:
            user_id: User UUID

        Returns:
            ExcelUploadSession if found, None otherwise
        """
        statement = select(ExcelUploadSession).where(
            ExcelUploadSession.user_id == user_id,
            ExcelUploadSession.processing_status == ExcelUploadProcessingStatus.PROCESSING
        )
        return self.db.exec(statement).first()

    def update_excel_with_status(self, file_path: str, invoice_updates: list[dict]) -> None:
        """
        Update Excel file with invoice status and reason.

        Args:
            file_path: Path to Excel file
            invoice_updates: List of dicts with invoice_number, status, reason

        Raises:
            ValueError: If Excel file cannot be updated
        """
        try:
            # Read existing Excel file
            abs_path = Path.cwd() / file_path
            df = pd.read_excel(abs_path, engine='openpyxl')

            # Create lookup dictionary for updates
            updates_dict = {
                update['invoice_number']: {
                    'status': update['status'],
                    'reason': update.get('reason', '')
                }
                for update in invoice_updates
            }

            # Update status and reason columns
            for idx, row in df.iterrows():
                invoice_num = str(row['invoice_number']).strip()
                if invoice_num in updates_dict:
                    df.at[idx, 'status'] = updates_dict[invoice_num]['status']
                    df.at[idx, 'reason'] = updates_dict[invoice_num]['reason']

            # Write back to Excel file
            with pd.ExcelWriter(abs_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Invoices')

                # Preserve column widths
                worksheet = writer.sheets['Invoices']
                column_widths = {
                    'A': 15, 'B': 25, 'C': 30, 'D': 12, 'E': 10,
                    'F': 15, 'G': 15, 'H': 15, 'I': 30,
                }
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width

        except Exception as e:
            raise ValueError(f"Error updating Excel file: {str(e)}")

    def generate_excel_from_database(self, invoices: list) -> BytesIO:
        """
        Generate Excel file from database invoice records.
        Used for exporting processed invoices with their status and results.

        Args:
            invoices: List of AutomationInvoice objects

        Returns:
            BytesIO object containing Excel file
        """
        # Prepare data rows
        rows = []
        for invoice in invoices:
            invoice_data = invoice.invoice_data

            # Extract first item (single item per invoice in current implementation)
            item = invoice_data.get('items', [{}])[0] if invoice_data.get('items') else {}

            row = {
                # Invoice identification — invoice_number is not stored in
                # invoice_data JSON anymore; the assigned value lives on the
                # AutomationInvoice column (set at transfer time)
                "invoice_number": invoice.invoice_number or '',
                "invoice_type": invoice_data.get('invoice_type', ''),
                "invoice_date": invoice_data.get('invoice_date', ''),

                # Buyer information
                "buyer_ntn_cnic": invoice_data.get('buyer_ntn_cnic', ''),
                "buyer_business_name": invoice_data.get('buyer_business_name', ''),
                "buyer_province": invoice_data.get('buyer_province', ''),
                "buyer_address": invoice_data.get('buyer_address', ''),
                "buyer_registration_type": invoice_data.get('buyer_registration_type', ''),

                # Item details - note: saved_item_code not stored in invoice_data, so we show empty
                "saved_item_code": "",
                "quantity": item.get('quantity', 0),
                "value_sales_excluding_st": item.get('value_sales_excluding_st', 0),
                "fixed_notified_value_or_retail_price": item.get('fixed_notified_value_or_retail_price', 0),
                "further_tax": item.get('further_tax', 0),
                "discount": item.get('discount', 0),
                "income_tax": invoice_data.get('income_tax', '236G'),
                "withholding_tax_amount": item.get('withholding_tax_amount', 0),

                # Scheduling
                "scheduled_date": invoice.scheduled_date.isoformat() if invoice.scheduled_date else '',
                "scheduled_time": invoice.scheduled_time.strftime('%H:%M') if invoice.scheduled_time else '',
            }
            rows.append(row)

        # Create DataFrame — TEMPLATE_COLUMNS no longer includes invoice_number,
        # but the exported file keeps it as the leading column for users
        df = pd.DataFrame(rows, columns=["invoice_number"] + self.TEMPLATE_COLUMNS)

        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Invoices')

            # Get workbook and worksheet
            worksheet = writer.sheets['Invoices']

            # Set column widths (same as template)
            from openpyxl.utils import get_column_letter

            column_widths = [
                15, 15, 12, 15, 25, 15, 30, 20,
                20, 10, 20, 25, 12, 12, 12, 20, 15, 15
            ]

            for idx, width in enumerate(column_widths, start=1):
                col_letter = get_column_letter(idx)
                worksheet.column_dimensions[col_letter].width = width

        output.seek(0)
        return output

    def _format_reason(self, invoice) -> str:
        """
        Format reason text based on invoice status and results.

        Args:
            invoice: AutomationInvoice object

        Returns:
            Formatted reason string
        """
        if invoice.status.value == 'transferred':
            reason = "Successfully transferred to main database"
            if invoice.fbr_response and invoice.fbr_response.get('reference_number'):
                reason += f" (Ref: {invoice.fbr_response['reference_number']})"
            return reason
        elif invoice.status.value == 'transfer_failed':
            return invoice.transfer_error or "Transfer to main database failed"
        elif invoice.status.value == 'failed':
            return invoice.validation_errors or "Processing failed"
        elif invoice.status.value == 'expired':
            return "Scheduled time is in the past"
        elif invoice.status.value == 'pending':
            return "Waiting for scheduled time"
        elif invoice.status.value == 'validated':
            return "Validated, awaiting transfer"
        else:
            return ""
