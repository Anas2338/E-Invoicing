"""
Manual Excel Helper - standalone functions for manual (non-automation) Excel operations.
Extracted from ExcelService to remove the automation DB dependency.
"""

from typing import Optional
from datetime import date, datetime
import pandas as pd
from io import BytesIO
from uuid import UUID
from sqlmodel import Session, select
import logging

from src.utils.excel_validator import ExcelValidator
from src.schemas.invoice import Environment


def _clean_ntn_cnic(value) -> str:
    """Normalize NTN/CNIC value from Excel cell.

    Returns empty string for blank cells and common placeholder values
    that Excel/pandas produce (e.g., 0, 0.0, nan). Handles float-to-int
    conversion for numeric cells (e.g., 1234567.0 → "1234567").
    """
    if pd.isna(value):
        return ""
    if isinstance(value, float):
        if value == int(value):
            value = int(value)
    s = str(value).strip()
    if s in ("", "0", "0.0", "nan", "None", "none", "null", "N/A", "n/a", "-", "NA", "na", "Nil", "nil"):
        return ""
    if s.endswith(".0") and len(s) > 2:
        s = s[:-2]
    return s

logger = logging.getLogger(__name__)

# pandas' default NA values list with "None" removed. pandas treats the
# literal string "None" as a missing value by default, which silently
# swallowed the income_tax="None" option from Excel files (cells came back
# as NaN and fell back to "236G"). Reading with this list keeps "None" intact
# while preserving the same blank/junk-cell behavior everywhere else.
_EXCEL_NA_VALUES = [
    "", "#N/A", "#N/A N/A", "#NA", "-1.#IND", "-1.#QNAN", "-NaN", "-nan",
    "1.#IND", "1.#QNAN", "<NA>", "N/A", "NA", "NULL", "NaN",
    "n/a", "nan", "null",
]


def _excel_float(value, default: float = 0.0) -> float:
    """Parse an Excel cell to float, falling back to `default` for blank cells."""
    if not pd.notna(value):
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def _excel_float_or_none(value) -> float | None:
    """Parse an Excel cell to float; None for blank or non-numeric cells."""
    if not pd.notna(value):
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _withholding_tax_rate(income_tax: str) -> float:
    """Withholding tax rate for an income tax type.

    236H -> 0.5%, 236G -> 0.1%, None (no income tax) -> 0.
    """
    if income_tax == "236H":
        return 0.005
    if income_tax == "None":
        return 0.0
    return 0.001


def _get_user_invoice_settings(user) -> tuple[str, int, int, bool]:
    """Return (prefix, start_number, padding, include_year) for a user."""
    return (
        user.invoice_prefix or "INV-",
        user.invoice_start_number or 1,
        user.invoice_padding or 4,
        user.invoice_include_year or False,
    )


def _generate_auto_invoice_numbers(
    db: Session,
    user,
    count: int,
    taken: set[str],
) -> list[str]:
    """Generate `count` sequential auto invoice numbers for a user.

    Mirrors the /profile/next-invoice-number logic: the sequence starts at
    the latest invoice's trailing number + 1 (or the configured start number)
    and skips any numbers already present in `taken` (existing invoices and
    numbers explicitly provided elsewhere in the same file).
    """
    from src.utils.helpers import (
        extract_invoice_number_suffix,
        format_invoice_number,
    )
    from src.models.invoice import Invoice as InvoiceModel
    from sqlmodel import select

    prefix, start_number, padding, include_year = _get_user_invoice_settings(user)

    latest = db.exec(
        select(InvoiceModel)
        .where(
            InvoiceModel.user_id == user.id,
            InvoiceModel.is_deleted == False,
        )
        .order_by(InvoiceModel.created_at.desc())
    ).first()

    if latest and latest.external_id:
        suffix = extract_invoice_number_suffix(latest.external_id)
        next_number = suffix + 1 if suffix is not None else start_number
    else:
        next_number = start_number

    numbers: list[str] = []
    for _ in range(count):
        candidate = format_invoice_number(prefix, next_number, padding, include_year)
        while candidate in taken:
            next_number += 1
            candidate = format_invoice_number(prefix, next_number, padding, include_year)
        taken.add(candidate)
        numbers.append(candidate)
        next_number += 1
    return numbers


MANUAL_TEMPLATE_COLUMNS = [
    "invoice_number",
    "invoice_type",
    "invoice_date",
    "buyer_ntn_cnic",
    "buyer_business_name",
    "buyer_province",
    "buyer_address",
    "buyer_registration_type",
    "saved_item_code",
    "product_description",
    "quantity",
    "value_sales_excluding_st",
    "fixed_notified_value_or_retail_price",
    "further_tax",
    "discount",
    "income_tax",
    "withholding_tax_amount",
]


def _shade_columns_to_last_row(
    buf: BytesIO,
    all_column_letters: list[str],
    fill_column_letters: list[str],
) -> None:
    """Apply per-column styles for the full column height (rows 1..1048576).

    Excel's ``<col style="...">`` attribute applies the style to every cell in
    that column, which styles the full column height without materializing
    millions of cells. openpyxl does not serialize column styles, so the style
    indices are resolved from styles.xml and injected into the sheet's <col>
    elements after saving.

    Every column gets a thin border; the ``fill_column_letters`` additionally
    get the light blue background. The column styles use plain fonts and no
    alignment so data rows aren't affected by header formatting.
    """
    import re
    import zipfile
    from openpyxl.utils import column_index_from_string

    rgb = '00DDEBF7'
    fill_numbers = {column_index_from_string(letter) for letter in fill_column_letters}
    all_numbers = [column_index_from_string(letter) for letter in all_column_letters]

    with zipfile.ZipFile(BytesIO(buf.getvalue())) as zin:
        styles_xml = zin.read('xl/styles.xml').decode('utf-8')

        # Resolve the fill index in the workbook's style table
        # (fills[0] and fills[1] are the mandatory defaults).
        fills_match = re.search(r'<fills[^>]*>(.*?)</fills>', styles_xml, re.S)
        fill_index = None
        for idx, fill in enumerate(re.finditer(r'<fill>(.*?)</fill>', fills_match.group(1), re.S)):
            if f'rgb="{rgb}"' in fill.group(1):
                fill_index = idx
                break
        if fill_index is None:
            raise RuntimeError('light blue fill not found in workbook styles')

        # Resolve the thin border index (a border with style="thin" sides).
        borders_match = re.search(r'<borders[^>]*>(.*?)</borders>', styles_xml, re.S)
        border_pattern = re.compile(r'<border\b[^>]*?/>|<border\b.*?</border>', re.S)
        thin_border_id = None
        for idx, border_xml in enumerate(border_pattern.findall(borders_match.group(1))):
            if 'style="thin"' in border_xml:
                thin_border_id = idx
                break
        if thin_border_id is None:
            raise RuntimeError('thin border not found in workbook styles')

        # Which fonts are bold? Column styles must not reference them.
        # Match both full <font>…</font> elements and self-closing <font />.
        fonts_match = re.search(r'<fonts[^>]*>(.*?)</fonts>', styles_xml, re.S)
        fonts_bold = []
        if fonts_match:
            font_pattern = re.compile(r'<font\b[^>]*?/>|<font\b.*?</font>', re.S)
            for font_xml in font_pattern.findall(fonts_match.group(1)):
                fonts_bold.append(bool(re.search(r'<b\b', font_xml)))

        # Find plain (non-bold, no alignment) cell styles for the exact
        # fill/border combinations — rows 3+ inherit these via <col>.
        cell_xfs_match = re.search(r'<cellXfs[^>]*>(.*?)</cellXfs>', styles_xml, re.S)
        xfs = [m.group(0) for m in re.finditer(r'<xf\b[^>]*?>', cell_xfs_match.group(1), re.S)]

        def find_style_xf(fill_id: int, border_id: int) -> int | None:
            for xf_idx, xf in enumerate(xfs):
                f_id = int(re.search(r'fillId="(\d+)"', xf).group(1))
                b_id = int(re.search(r'borderId="(\d+)"', xf).group(1))
                font_id = int(re.search(r'fontId="(\d+)"', xf).group(1))
                if (
                    f_id == fill_id
                    and b_id == border_id
                    and not fonts_bold[font_id]
                    and 'applyAlignment' not in xf
                ):
                    return xf_idx
            return None

        fill_border_style = find_style_xf(fill_index, thin_border_id)
        border_only_style = find_style_xf(0, thin_border_id)
        if fill_border_style is None or border_only_style is None:
            raise RuntimeError('plain (fill/border) cell styles not found for shading')

        style_by_column = {
            n: (fill_border_style if n in fill_numbers else border_only_style)
            for n in all_numbers
        }

        def patch_cols(sheet_xml: str) -> str:
            def add_style(match: re.Match) -> str:
                attrs = match.group(1)
                m_num = re.search(r'min="(\d+)"', attrs)
                style = style_by_column[int(m_num.group(1))]
                if 'style=' in attrs:
                    attrs = re.sub(r'style="[^"]*"', f'style="{style}"', attrs)
                else:
                    attrs += f' style="{style}"'
                return f'<col{attrs}/>'

            patched = sheet_xml
            for n in all_numbers:
                patched = re.sub(r'<col\b([^>]*?)/>', add_style, patched)
            return patched

        entries: list[tuple[zipfile.ZipInfo, bytes]] = []
        for item in zin.infolist():
            content = zin.read(item.filename)
            if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                content = patch_cols(content.decode('utf-8')).encode('utf-8')
            entries.append((item, content))

    buf.seek(0)
    buf.truncate()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item, content in entries:
            zout.writestr(item, content)


def generate_manual_excel_template(
    invoice_types: list[str] | None = None,
    provinces: list[str] | None = None,
) -> BytesIO:
    """Generate Excel template for manual invoice upload with dropdown validations.

    Args:
        invoice_types: Valid invoice type names (e.g. ['Sale Invoice', 'Debit Note']).
                       Falls back to default list if None.
        provinces: Valid province names (e.g. ['PUNJAB', 'SINDH', ...]).
                   Falls back to default list if None.
    """
    if not invoice_types:
        invoice_types = ["Sale Invoice", "Debit Note"]
    if not provinces:
        provinces = [
            "PUNJAB", "SINDH", "KPK", "BALOCHISTAN",
            "ISLAMABAD", "GILGIT BALTISTAN", "AZAD JAMMU KASHMIR"
        ]

    df = pd.DataFrame(columns=MANUAL_TEMPLATE_COLUMNS)

    sample_row = pd.DataFrame([{
        "invoice_number": "INV-001",
        "invoice_type": "Sale Invoice",
        "invoice_date": "2026-05-12",
        "buyer_ntn_cnic": "1234567",
        "buyer_business_name": "ABC Corporation",
        "buyer_province": "PUNJAB",
        "buyer_address": "123 Main Street, Lahore",
        "buyer_registration_type": "Registered",
        "saved_item_code": "ITEM001",
        "product_description": "Laptop Computer",
        "quantity": "2",
        "value_sales_excluding_st": "50000",
        "fixed_notified_value_or_retail_price": "50000",
        "further_tax": "0",
        "discount": "0",
        "income_tax": "236G",
        "withholding_tax_amount": "50",
    }])
    df = pd.concat([df, sample_row], ignore_index=True)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Invoices')

        workbook = writer.book
        worksheet = writer.sheets['Invoices']

        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        # — Column widths (auto-fit based on header text) —
        column_headers = MANUAL_TEMPLATE_COLUMNS
        for idx, header in enumerate(column_headers, start=1):
            col_letter = get_column_letter(idx)
            # Calculate width: header length + padding, with min/max bounds
            width = max(len(header) + 3, 12)
            width = min(width, 40)
            worksheet.column_dimensions[col_letter].width = width

        # — Freeze the header row so it stays visible when scrolling —
        worksheet.freeze_panes = 'A2'

        # — Data validation dropdowns (inline comma-separated lists) —
        last_data_row = 1048576  # covers entire column (Excel max rows)

        # Column letters derived from MANUAL_TEMPLATE_COLUMNS so they stay
        # correct if columns are added/removed.
        def col_letter(column_name: str) -> str:
            return get_column_letter(MANUAL_TEMPLATE_COLUMNS.index(column_name) + 1)

        option_sets: dict[str, list[str]] = {
            col_letter('invoice_type'): invoice_types,                          # invoice_type
            col_letter('buyer_province'): provinces,                            # buyer_province
            col_letter('buyer_registration_type'): [
                "Registered", "Unregistered", "Final Consumer",
            ],                                                                  # buyer_registration_type
            col_letter('income_tax'): ["None", "236G", "236H"],                 # income_tax
        }

        for letter, values in option_sets.items():
            options_str = ','.join(values)
            dv = DataValidation(
                type='list',
                formula1=f'"{options_str}"',
                allow_blank=True,
            )
            dv.error = 'Please select a valid value from the dropdown list.'
            dv.errorTitle = 'Invalid value'
            dv.prompt = 'Select from the dropdown list'
            dv.promptTitle = 'Valid options'
            worksheet.add_data_validation(dv)
            dv.add(f'{letter}2:{letter}{last_data_row}')

        # — Light blue fill + thin borders across the full sheet —
        from openpyxl.styles import PatternFill, Font, Border, Side

        # Bold headings only in the header row — data rows stay plain via
        # the column styles patched below.
        header_font = Font(bold=True)
        thin_side = Side(style='thin')
        thin_border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
        )
        for cell in worksheet[1]:
            cell.font = header_font
            cell.border = thin_border
        # Register the thin-border style on the demo row too, so the column
        # patch below can reference it for rows 3+ (full column height).
        for cell in worksheet[2]:
            cell.border = thin_border

        light_blue_fill = PatternFill(
            start_color='DDEBF7', end_color='DDEBF7', fill_type='solid'
        )
        highlighted_columns = {
            "invoice_type", "invoice_date", "buyer_business_name",
            "buyer_province", "buyer_address", "buyer_registration_type",
            "saved_item_code", "product_description", "quantity",
            "value_sales_excluding_st",
        }
        highlighted_letters = sorted(
            col_letter(column_name) for column_name in highlighted_columns
        )
        for column_name in highlighted_columns:
            letter = col_letter(column_name)
            # Fill the header cell so openpyxl registers the style, then patch
            # the <col> elements below so Excel shades the full column height.
            worksheet[f'{letter}1'].fill = light_blue_fill
            # Fill the demo row cell explicitly so the sample line shows the
            # colour in every viewer (rows 3+ rely on the <col> style patch).
            worksheet[f'{letter}2'].fill = light_blue_fill

        all_column_letters = [
            get_column_letter(i) for i in range(1, len(MANUAL_TEMPLATE_COLUMNS) + 1)
        ]

    _shade_columns_to_last_row(output, all_column_letters, highlighted_letters)

    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# Staging parser functions (non-failing, per-field error tracking)
# ---------------------------------------------------------------------------


def _validate_staging_row(
    row: dict,
    saved_items_dict: dict,
    today: date,
) -> dict[str, list[str]]:
    """Validate a single staging row, returning per-field errors.

    Pure function — does not modify any state. Returns a dict mapping
    field names to lists of error messages. Empty dict = row is valid.
    """
    errors: dict[str, list[str]] = {}

    invoice_number = str(row.get("invoice_number", "")).strip()
    if not invoice_number:
        errors["invoice_number"] = ["invoice_number is required"]
        # Can't validate most fields without an invoice number
        return errors

    # --- invoice_date ---
    invoice_date_str = str(row.get("invoice_date", "")).strip()
    if not invoice_date_str:
        errors["invoice_date"] = ["invoice_date is required"]
    else:
        try:
            from datetime import datetime as dt
            parsed = dt.strptime(invoice_date_str, "%Y-%m-%d").date()
            if parsed > today:
                errors["invoice_date"] = [
                    f"invoice_date '{invoice_date_str}' is in the future"
                ]
        except ValueError:
            errors["invoice_date"] = [
                f"invalid invoice_date format '{invoice_date_str}'"
            ]

    # --- saved_item_code ---
    saved_item_code = str(row.get("saved_item_code", "")).strip()
    if not saved_item_code:
        errors["saved_item_code"] = ["saved_item_code is required"]
    else:
        saved_item = saved_items_dict.get(saved_item_code)
        if saved_item is None:
            errors["saved_item_code"] = [
                f"'{saved_item_code}' not found in your saved items"
            ]

    # --- buyer_registration_type ---
    buyer_reg_type = str(row.get("buyer_registration_type", "")).strip()
    if not buyer_reg_type:
        errors["buyer_registration_type"] = ["buyer registration type is required"]
    elif buyer_reg_type not in ("Registered", "Unregistered", "Final Consumer"):
        errors["buyer_registration_type"] = [
            f"'{buyer_reg_type}' is not a valid registration type"
        ]

    # --- buyer_ntn_cnic (required for registered buyers) ---
    buyer_ntn = str(row.get("buyer_ntn_cnic", "")).strip()
    if buyer_reg_type == "Registered" and not buyer_ntn:
        errors["buyer_ntn_cnic"] = [
            "buyer NTN/CNIC is required for registered buyers"
        ]

    # --- buyer_business_name ---
    buyer_name = str(row.get("buyer_business_name", "")).strip()
    if not buyer_name:
        errors["buyer_business_name"] = ["buyer business name is required"]

    # --- buyer_province ---
    buyer_province = str(row.get("buyer_province", "")).strip()
    if not buyer_province:
        errors["buyer_province"] = ["buyer province is required"]

    # --- buyer_address ---
    buyer_address = str(row.get("buyer_address", "")).strip()
    if not buyer_address:
        errors["buyer_address"] = ["buyer address is required"]

    # --- income_tax ---
    income_tax = str(row.get("income_tax", "")).strip()
    if income_tax and income_tax not in ("236G", "236H", "None"):
        errors["income_tax"] = [f"income_tax '{income_tax}' is invalid"]

    # --- quantity ---
    try:
        quantity = float(row.get("quantity", 0))
        if quantity <= 0:
            errors["quantity"] = ["quantity must be greater than 0"]
    except (ValueError, TypeError):
        errors["quantity"] = ["quantity must be a valid number"]

    # --- value_sales_excluding_st ---
    try:
        value = float(row.get("value_sales_excluding_st", 0))
        if value <= 0:
            errors["value_sales_excluding_st"] = [
                "value_sales_excluding_st must be greater than 0"
            ]
    except (ValueError, TypeError):
        errors["value_sales_excluding_st"] = [
            "value_sales_excluding_st must be a valid number"
        ]

    # --- discount (if provided) must not exceed value ---
    try:
        discount = float(row.get("discount", 0))
        if discount > 0 and discount > value:
            errors["discount"] = [
                f"discount ({discount}) cannot exceed "
                f"value_sales_excluding_st ({value})"
            ]
    except (ValueError, TypeError):
        pass

    # --- fixed_notified_value check for 3rd Schedule Goods ---
    if saved_item_code and saved_items_dict.get(saved_item_code):
        item = saved_items_dict[saved_item_code]
        if getattr(item, "transaction_type", "") == "3rd Schedule Goods":
            try:
                fixed_val = float(
                    row.get("fixed_notified_value_or_retail_price", 0)
                )
                if fixed_val < value:
                    errors["fixed_notified_value_or_retail_price"] = [
                        f"fixed_notified_value_or_retail_price ({fixed_val}) "
                        f"must be >= value_sales_excluding_st ({value})"
                    ]
            except (ValueError, TypeError):
                pass

    return errors


def _validate_staging_row_full(
    row: dict,
    saved_items_dict: dict,
    today: date,
    other_invoice_numbers: set[str] | None = None,
    existing_invoice_numbers: set[str] | None = None,
) -> dict[str, list[str]]:
    """Validate a staging row against ALL rules, including cross-row checks.

    Runs the per-field validation from _validate_staging_row, then adds
    duplicate invoice number checks:
      - other_invoice_numbers: numbers used by other rows in the same
        session/file
      - existing_invoice_numbers: numbers already saved in the user's
        invoice history

    Pure function — does not modify any state. Returns a dict mapping
    field names to lists of error messages. Empty dict = row is valid.
    """
    errors = _validate_staging_row(row, saved_items_dict, today)

    invoice_number = str(row.get("invoice_number", "")).strip()
    if invoice_number and not errors.get("invoice_number"):
        if other_invoice_numbers and invoice_number in other_invoice_numbers:
            errors["invoice_number"] = [
                f"duplicate invoice number '{invoice_number}' within the same Excel file"
            ]
        elif existing_invoice_numbers and invoice_number in existing_invoice_numbers:
            errors["invoice_number"] = [
                f"invoice number '{invoice_number}' already exists in your invoice history"
            ]

    return errors


def _compute_staging_fields(
    row: dict,
    saved_items_dict: dict,
    seller_info: dict,
    income_tax: str,
) -> dict:
    """Compute derived fields for a staging row.

    These include financial calculations and saved-item lookups.
    Returns a dict of computed fields that can be merged into the row.
    """
    computed: dict = {}

    saved_item_code = str(row.get("saved_item_code", "")).strip()
    saved_item = saved_items_dict.get(saved_item_code)

    # --- Resolve saved item fields ---
    if saved_item:
        # Template product_description wins if provided; fall back to the
        # saved item's description when the cell is blank.
        template_desc = str(row.get("product_description") or "").strip()
        computed["product_description"] = (
            template_desc or saved_item.product_description
        )
        computed["hs_code"] = saved_item.hs_code
        computed["rate"] = str(saved_item.default_rate or "18")
        computed["uom"] = saved_item.default_uom or "NOS"
        computed["sale_type"] = saved_item.transaction_type or ""
        computed["transaction_type_id"] = saved_item.transaction_type or ""
        computed["sro_schedule_no"] = saved_item.sro_schedule_no or ""
        computed["sro_item_serial_no"] = saved_item.sro_item_serial_no or ""

    # --- Financial calculations ---
    try:
        quantity = float(row.get("quantity", 0))
    except (ValueError, TypeError):
        quantity = 0

    try:
        value_sales_excluding_st = float(row.get("value_sales_excluding_st", 0))
    except (ValueError, TypeError):
        value_sales_excluding_st = 0

    try:
        fixed_notified_value = float(
            row.get("fixed_notified_value_or_retail_price", 0)
        )
    except (ValueError, TypeError):
        fixed_notified_value = 0

    try:
        further_tax = float(row.get("further_tax", 0))
    except (ValueError, TypeError):
        further_tax = 0

    try:
        discount = float(row.get("discount", 0))
    except (ValueError, TypeError):
        discount = 0

    tax_rate = float(saved_item.default_rate) if saved_item and saved_item.default_rate else 18.0
    base_value = max(value_sales_excluding_st, fixed_notified_value)
    computed["sales_tax_applicable"] = round((base_value * tax_rate) / 100, 2)
    computed["total_values"] = round(
        base_value + computed["sales_tax_applicable"] + further_tax - discount, 2
    )

    # Withholding tax (0.1% for 236G, 0.5% for 236H, 0 for None)
    wht_rate = _withholding_tax_rate(income_tax)
    wht = row.get("withholding_tax_amount")
    if wht is not None:
        try:
            computed["withholding_tax_amount"] = float(wht)
        except (ValueError, TypeError):
            computed["withholding_tax_amount"] = round(value_sales_excluding_st * wht_rate, 2)
    else:
        computed["withholding_tax_amount"] = round(value_sales_excluding_st * wht_rate, 2)

    computed["sales_tax_withheld_at_source"] = 0
    computed["extra_tax"] = 0
    computed["fed_payable"] = 0
    computed["item_rate"] = (
        round(value_sales_excluding_st / quantity, 2)
        if quantity > 0 else None
    )

    # Seller info
    computed.update(seller_info)

    return computed


def parse_excel_for_staging(
    file_source: BytesIO,
    user_id: UUID,
    db: Session,
    automation_invoice_numbers: set[str] | None = None,
) -> list[dict]:
    """Parse an Excel file for staging, returning ALL rows without failing.

    Each row dict includes:
      - All 17 template fields (as parsed from Excel)
      - Computed fields (from saved item lookup + financial calculations)
      - Seller fields (from user profile)
      - Validation state: is_valid, field_errors
      - Metadata: excel_row_number, group_key (invoice_number)

    Blank invoice_number cells are auto-issued the user's next sequential
    invoice number (based on their numbering settings and last invoice).

    Never raises ValueError. Returns empty list if no valid invoice data.
    Skips the sample row (INV-001).
    """
    from src.models.user import User
    from src.models.user_saved_product import UserSavedProduct

    # --- Parse the Excel file ---
    try:
        df = pd.read_excel(
            file_source,
            engine="openpyxl",
            dtype={"saved_item_code": str, "income_tax": str},
            keep_default_na=False,
            na_values=_EXCEL_NA_VALUES,
        )
    except MemoryError:
        return []

    # --- Normalize column names ---
    # Handle column names with __N suffixes (from pandas duplicate column handling)
    # Also handle case-insensitive matches and strip whitespace
    template_set = set(MANUAL_TEMPLATE_COLUMNS)
    # Build a mapping: normalize Excel column names to our template names
    import re
    col_rename = {}
    for col in df.columns:
        col_str = str(col).strip()
        # Strip __N suffixes first
        clean = re.sub(r"__\d+$", "", col_str)
        # Try direct match first, then case-insensitive match
        if clean in template_set:
            col_rename[col] = clean
        else:
            # Try lowercase matching
            for tc in template_set:
                if clean.lower() == tc.lower():
                    col_rename[col] = tc
                    break
    # Apply rename and keep only matched columns
    if col_rename:
        df = df.rename(columns=col_rename)
    cols_to_keep = [c for c in df.columns if c in template_set]
    unmatched = [c for c in df.columns if c not in template_set]
    if unmatched:
        logger.info("Dropping unmatched columns from Excel: %s", unmatched)
    df = df[cols_to_keep]

    # Drop fully-empty rows (invoice_number is optional — blank rows get an
    # auto-issued number below) and the sample row
    df = df.dropna(how="all")
    if df.empty:
        return []
    df = df[df["invoice_number"].astype(str).str.strip() != "INV-001"]
    if df.empty:
        return []

    # --- Check for duplicate invoice numbers already in the database ---
    from src.models.invoice import Invoice as InvoiceModel
    excel_invoice_numbers: set[str] = set()
    for _, raw_row in df.iterrows():
        inv_num = _clean_ntn_cnic(raw_row.get("invoice_number"))
        if inv_num:
            excel_invoice_numbers.add(inv_num)

    existing_invoice_numbers: set[str] = set()
    if excel_invoice_numbers:
        existing = db.exec(
            select(InvoiceModel.external_id).where(
                InvoiceModel.external_id.in_(excel_invoice_numbers),
                InvoiceModel.user_id == user_id,
                InvoiceModel.is_deleted == False,
            )
        ).all()
        existing_invoice_numbers = set(existing)

    # --- Fetch seller info ---
    seller_info: dict = {}
    user = db.get(User, user_id)
    if user:
        seller_info = {
            "seller_ntn_cnic": user.fbr_seller_ntn or "",
            "seller_business_name": user.fbr_business_name or "",
            "seller_province": user.fbr_seller_province or "",
            "seller_address": user.fbr_seller_address or "",
        }

    # --- Auto-issue invoice numbers for blank rows ---
    # One new sequential number per blank row, based on the user's numbering
    # settings and latest invoice. Skips numbers already present in the file,
    # in the user's invoice history, or in the automation database (not yet
    # transferred to the main database).
    auto_numbers: list[str] = []
    if user:
        taken = (
            set(excel_invoice_numbers)
            | set(existing_invoice_numbers)
            | (automation_invoice_numbers or set())
        )
        blank_count = sum(
            1 for _, raw_row in df.iterrows()
            if not _clean_ntn_cnic(raw_row.get("invoice_number"))
        )
        if blank_count:
            auto_numbers = _generate_auto_invoice_numbers(
                db, user, blank_count, taken
            )
    auto_numbers_iter = iter(auto_numbers)

    # --- Fetch saved items ---
    statement = select(UserSavedProduct).where(
        UserSavedProduct.user_id == user_id,
        UserSavedProduct.is_active == 1,
    )
    saved_items = db.exec(statement).all()
    saved_items_dict = {item.item_code: item for item in saved_items}

    # --- Process each row ---
    today = date.today()
    rows: list[dict] = []
    seen_invoice_numbers: set[str] = set()       # duplicates within same file
    invoice_dates: dict[str, str] = {}            # first seen date per invoice number

    for row_idx, raw_row in df.iterrows():
        excel_row_number = row_idx + 2  # 1-based + header

        invoice_number = _clean_ntn_cnic(raw_row.get("invoice_number"))
        if not invoice_number:
            # Blank invoice_number -> auto-issue the next number for this user
            try:
                invoice_number = next(auto_numbers_iter)
            except StopIteration:
                pass

        # Parse all 17 template fields with defaults
        row_data: dict = {
            "excel_row_number": excel_row_number,
            "invoice_number": invoice_number,
            "invoice_type": str(raw_row.get("invoice_type", "")).strip()
            if pd.notna(raw_row.get("invoice_type"))
            else "Sale Invoice",
            "invoice_date": pd.to_datetime(raw_row["invoice_date"]).strftime('%Y-%m-%d')
            if pd.notna(raw_row.get("invoice_date"))
            else "",
            "buyer_ntn_cnic": _clean_ntn_cnic(raw_row.get("buyer_ntn_cnic")),
            "buyer_business_name": str(raw_row.get("buyer_business_name", "")).strip()
            if pd.notna(raw_row.get("buyer_business_name"))
            else "",
            "buyer_province": str(raw_row.get("buyer_province", "")).strip()
            if pd.notna(raw_row.get("buyer_province"))
            else "",
            "buyer_address": str(raw_row.get("buyer_address", "")).strip()
            if pd.notna(raw_row.get("buyer_address"))
            else "",
            "buyer_registration_type": str(raw_row.get("buyer_registration_type", "")).strip()
            if pd.notna(raw_row.get("buyer_registration_type"))
            else "Registered",
            "saved_item_code": str(raw_row.get("saved_item_code", "")).strip()
            if pd.notna(raw_row.get("saved_item_code"))
            else "",
            "product_description": str(raw_row.get("product_description", "")).strip()
            if pd.notna(raw_row.get("product_description"))
            else "",
            "quantity": round(_excel_float(raw_row.get("quantity")), 2),
            "value_sales_excluding_st": round(
                _excel_float(raw_row.get("value_sales_excluding_st")), 2
            ),
            "fixed_notified_value_or_retail_price": round(
                _excel_float(raw_row.get("fixed_notified_value_or_retail_price")), 2
            ),
            "further_tax": round(_excel_float(raw_row.get("further_tax")), 0),
            "discount": round(_excel_float(raw_row.get("discount")), 2),
            "income_tax": str(raw_row.get("income_tax", "")).strip()
            if pd.notna(raw_row.get("income_tax"))
            else "236G",
            "withholding_tax_amount": _excel_float_or_none(
                raw_row.get("withholding_tax_amount")
            ),
        }

        # Determine income_tax for this row
        income_tax = row_data["income_tax"]
        income_tax_raw = str(raw_row.get("income_tax", "")).strip() if pd.notna(raw_row.get("income_tax")) else ""
        if income_tax_raw in ("236G", "236H"):
            income_tax = income_tax_raw
        elif income_tax_raw:
            income_tax = income_tax_raw  # will be flagged as error by validation

        # Validate
        field_errors = _validate_staging_row(row_data, saved_items_dict, today)

        # Check for duplicate invoice number already in the database
        # (including automation invoices not yet transferred)
        if (
            not field_errors.get("invoice_number")
            and (
                invoice_number in existing_invoice_numbers
                or invoice_number in (automation_invoice_numbers or set())
            )
        ):
            field_errors["invoice_number"] = [
                f"invoice number '{invoice_number}' already exists in your invoice history"
            ]

        # Check for duplicate invoice number within the same file
        if (
            not field_errors.get("invoice_number")
            and invoice_number in seen_invoice_numbers
        ):
            field_errors["invoice_number"] = [
                f"duplicate invoice number '{invoice_number}' within the same Excel file"
            ]
        seen_invoice_numbers.add(invoice_number)

        # Check that all rows with the same invoice_number have the same date
        current_date = row_data.get("invoice_date", "")
        if (
            not field_errors.get("invoice_date")
            and invoice_number in invoice_dates
            and current_date
            and current_date != invoice_dates[invoice_number]
        ):
            field_errors["invoice_date"] = [
                f"invoice_date '{current_date}' does not match "
                f"previous date '{invoice_dates[invoice_number]}' "
                f"for invoice '{invoice_number}'"
            ]
        if current_date and invoice_number not in invoice_dates:
            invoice_dates[invoice_number] = current_date

        is_valid = len(field_errors) == 0

        # Compute derived fields (even for invalid rows — user should see them)
        computed = _compute_staging_fields(
            row_data, saved_items_dict, seller_info, income_tax
        )
        row_data.update(computed)

        row_data["group_key"] = invoice_number
        row_data["is_valid"] = is_valid
        row_data["is_dirty"] = False
        row_data["field_errors"] = field_errors

        rows.append(row_data)

    return rows


def build_invoices_from_rows(
    valid_rows: list[dict],
    seller_info: dict | None = None,
) -> list[dict]:
    """Group valid staging rows by invoice_number and build invoice dicts.

    Each unique invoice_number becomes one invoice with multiple line items.
    This is the shared grouping logic used by both the old upload flow
    and the new staging commit flow.
    """
    invoice_groups: dict[str, dict] = {}
    seller = seller_info or {}

    for row in valid_rows:
        invoice_number = row.get("invoice_number", "").strip()
        if not invoice_number:
            continue

        if invoice_number not in invoice_groups:
            invoice_groups[invoice_number] = {
                "external_id": invoice_number,
                "invoice_type": row.get("invoice_type", "Sale Invoice"),
                "invoice_date": row.get("invoice_date", ""),
                "transaction_type_id": row.get("transaction_type_id", ""),
                "seller_ntn_cnic": seller.get("seller_ntn_cnic", row.get("seller_ntn_cnic", "")),
                "seller_business_name": seller.get("seller_business_name", row.get("seller_business_name", "")),
                "seller_province": seller.get("seller_province", row.get("seller_province", "")),
                "seller_address": seller.get("seller_address", row.get("seller_address", "")),
                "buyer_ntn_cnic": row.get("buyer_ntn_cnic", ""),
                "buyer_business_name": row.get("buyer_business_name", ""),
                "buyer_province": row.get("buyer_province", ""),
                "buyer_address": row.get("buyer_address", ""),
                "buyer_registration_type": row.get("buyer_registration_type", "Registered"),
                "invoice_ref_no": "",
                "scenario_id": "",
                "items": [],
                "environment": Environment.PRODUCTION,
                "income_tax": row.get("income_tax", "236G"),
            }

        item = {
            "hs_code": row.get("hs_code", ""),
            "product_description": row.get("product_description", ""),
            "rate": row.get("rate", "18"),
            "uom": row.get("uom", "NOS"),
            "quantity": row.get("quantity", 0),
            "total_values": row.get("total_values", 0),
            "value_sales_excluding_st": row.get("value_sales_excluding_st", 0),
            "fixed_notified_value_or_retail_price": row.get("fixed_notified_value_or_retail_price", 0),
            "sales_tax_applicable": row.get("sales_tax_applicable", 0),
            "sales_tax_withheld_at_source": row.get("sales_tax_withheld_at_source", 0),
            "extra_tax": row.get("extra_tax", 0),
            "further_tax": row.get("further_tax", 0),
            "sro_schedule_no": row.get("sro_schedule_no", ""),
            "fed_payable": row.get("fed_payable", 0),
            "discount": row.get("discount", 0),
            "sale_type": row.get("sale_type", "01"),
            "sro_item_serial_no": row.get("sro_item_serial_no", ""),
            "income_tax_type": row.get("income_tax", "236G"),
            "withholding_tax_amount": row.get("withholding_tax_amount", 0),
        }
        invoice_groups[invoice_number]["items"].append(item)

    return list(invoice_groups.values())


def parse_excel_for_manual_invoice(
    file_source: BytesIO | str,
    user_id: UUID = None,
    main_db: Session = None,
    automation_invoice_numbers: set[str] | None = None,
) -> list[dict]:
    """Parse Excel file for manual invoice creation.
    Validates invoice_date is today or previous date (no future dates).
    Does NOT include scheduling fields (automation-only).
    """
    from src.models.user import User
    from src.models.user_saved_product import UserSavedProduct
    from src.models.invoice import Invoice as InvoiceModel

    try:
        df = pd.read_excel(
            file_source,
            engine='openpyxl',
            dtype={'saved_item_code': str, 'income_tax': str},
            keep_default_na=False,
            na_values=_EXCEL_NA_VALUES,
        )
    except MemoryError:
        raise MemoryError(
            "File is too large to process in memory. "
            "Please reduce the file size or split into smaller batches (max 1,000 rows)."
        )

    # invoice_number is optional — blank rows get an auto-issued number below
    df = df.dropna(how="all")
    if not df.empty:
        df = df[df['invoice_number'].astype(str).str.strip() != 'INV-001']

    seller_info = {}
    user = None
    if user_id and main_db:
        user = main_db.get(User, user_id)
        if user:
            seller_info = {
                "seller_ntn_cnic": user.fbr_seller_ntn or "",
                "seller_business_name": user.fbr_business_name or "",
                "seller_province": user.fbr_seller_province or "",
                "seller_address": user.fbr_seller_address or "",
            }

    saved_items_dict = {}
    if user_id and main_db:
        statement = select(UserSavedProduct).where(
            UserSavedProduct.user_id == user_id,
            UserSavedProduct.is_active == 1,
        )
        saved_items = main_db.exec(statement).all()
        saved_items_dict = {item.item_code: item for item in saved_items}

    excel_invoice_numbers = set()
    for _, row in df.iterrows():
        inv_num = str(row['invoice_number']).strip() if pd.notna(row['invoice_number']) else ""
        if inv_num and inv_num != 'INV-001':
            excel_invoice_numbers.add(inv_num)

    existing_invoice_numbers = set()
    if excel_invoice_numbers and main_db:
        existing = main_db.exec(
            select(InvoiceModel.external_id).where(
                InvoiceModel.external_id.in_(excel_invoice_numbers),
                InvoiceModel.is_deleted == False,
            )
        ).all()
        existing_invoice_numbers = set(existing)

    # Auto-issue invoice numbers for blank rows (one sequential number per row,
    # skipping numbers already present in the file, in the user's history, or
    # in the automation database — not yet transferred to the main database)
    auto_numbers: list[str] = []
    if user and main_db:
        taken = (
            set(excel_invoice_numbers)
            | set(existing_invoice_numbers)
            | (automation_invoice_numbers or set())
        )
        blank_count = sum(
            1 for _, row in df.iterrows()
            if not _clean_ntn_cnic(row.get("invoice_number"))
        )
        if blank_count:
            auto_numbers = _generate_auto_invoice_numbers(
                main_db, user, blank_count, taken
            )
    auto_numbers_iter = iter(auto_numbers)

    today = date.today()
    invoice_groups: dict[str, dict] = {}
    seen_invoice_numbers: set[str] = set()  # track duplicates within the same file
    validation_errors = []

    for row_idx, row in df.iterrows():
        invoice_number = _clean_ntn_cnic(row.get('invoice_number'))
        if not invoice_number:
            try:
                invoice_number = next(auto_numbers_iter)
            except StopIteration:
                pass
        excel_row = row_idx + 2

        if (
            invoice_number in existing_invoice_numbers
            or invoice_number in (automation_invoice_numbers or set())
        ):
            validation_errors.append(
                f"Row {excel_row} (Invoice {invoice_number}): "
                f"invoice number already exists in your history."
            )
            continue

        if invoice_number in seen_invoice_numbers:
            validation_errors.append(
                f"Row {excel_row} (Invoice {invoice_number}): "
                f"duplicate invoice number within the same Excel file."
            )
            continue
        seen_invoice_numbers.add(invoice_number)

        saved_item_code = str(row['saved_item_code']).strip() if pd.notna(row['saved_item_code']) else ""
        if not saved_item_code:
            validation_errors.append(f"Row {excel_row} (Invoice {invoice_number}): saved_item_code is required")
            continue

        saved_item = saved_items_dict.get(saved_item_code)
        if not saved_item:
            validation_errors.append(
                f"Row {excel_row} (Invoice {invoice_number}): "
                f"saved_item_code '{saved_item_code}' not found in your saved items"
            )
            continue

        invoice_date_str = ""
        if pd.notna(row['invoice_date']):
            try:
                invoice_date_parsed = pd.to_datetime(row['invoice_date'])
                invoice_date_str = invoice_date_parsed.strftime('%Y-%m-%d')
                invoice_date_obj = invoice_date_parsed.date()
                if invoice_date_obj > today:
                    validation_errors.append(
                        f"Row {excel_row} (Invoice {invoice_number}): "
                        f"invoice_date '{invoice_date_str}' is in the future."
                    )
                    continue
            except Exception:
                invoice_date_str = str(row['invoice_date']).strip()
        else:
            validation_errors.append(f"Row {excel_row} (Invoice {invoice_number}): invoice_date is required")
            continue

        income_tax = "236G"
        if pd.notna(row.get('income_tax')):
            income_tax_raw = str(row['income_tax']).strip()
            if income_tax_raw in ("236G", "236H", "None"):
                income_tax = income_tax_raw
            elif income_tax_raw:
                validation_errors.append(
                    f"Row {excel_row} (Invoice {invoice_number}): "
                    f"income_tax '{income_tax_raw}' is invalid."
                )
                continue

        quantity = round(_excel_float(row.get('quantity')), 2)
        value_sales_excluding_st = round(_excel_float(row.get('value_sales_excluding_st')), 2)
        fixed_notified_value_or_retail_price = round(
            _excel_float(row.get('fixed_notified_value_or_retail_price')), 2
        )
        further_tax = round(_excel_float(row.get('further_tax')), 0)
        discount = round(_excel_float(row.get('discount')), 2)

        # Parse withholding_tax_amount from Excel (optional, auto-calc if omitted)
        # WHT rate: 0.1% for 236G, 0.5% for 236H, 0 for None
        withholding_tax_amount = None
        if pd.notna(row.get('withholding_tax_amount')):
            try:
                withholding_tax_amount = float(row['withholding_tax_amount'])
            except (ValueError, TypeError):
                withholding_tax_amount = None
        if withholding_tax_amount is None:
            wht_rate = _withholding_tax_rate(income_tax)
            withholding_tax_amount = round(value_sales_excluding_st * wht_rate, 2)

        tax_rate = float(saved_item.default_rate) if saved_item.default_rate else 18.0
        base_value = max(value_sales_excluding_st, fixed_notified_value_or_retail_price)
        sales_tax_applicable = round((base_value * tax_rate) / 100, 2)
        total_values = round(base_value + sales_tax_applicable + further_tax - discount, 2)

        uom_code = saved_item.default_uom or "NOS"

        # Template product_description wins if provided; fall back to the
        # saved item's description when the cell is blank.
        template_desc = str(row.get('product_description') or "").strip()

        item = {
            "hs_code": saved_item.hs_code,
            "product_description": template_desc or saved_item.product_description,
            "rate": str(saved_item.default_rate or "18"),
            "uom": uom_code,
            "quantity": quantity,
            "total_values": total_values,
            "value_sales_excluding_st": value_sales_excluding_st,
            "fixed_notified_value_or_retail_price": fixed_notified_value_or_retail_price,
            "sales_tax_applicable": sales_tax_applicable,
            "sales_tax_withheld_at_source": 0,
            "extra_tax": 0,
            "further_tax": further_tax,
            "sro_schedule_no": saved_item.sro_schedule_no or "",
            "fed_payable": 0,
            "discount": discount,
            "sale_type": saved_item.transaction_type or "01",
            "sro_item_serial_no": saved_item.sro_item_serial_no or "",
            # Internal fields (not sent to FBR)
            "income_tax_type": income_tax,
            "withholding_tax_amount": withholding_tax_amount,
        }

        buyer_ntn_cnic = _clean_ntn_cnic(row['buyer_ntn_cnic'])
        buyer_business_name = str(row['buyer_business_name']).strip() if pd.notna(row['buyer_business_name']) else ""
        buyer_province = str(row['buyer_province']).strip() if pd.notna(row['buyer_province']) else ""
        buyer_address = str(row['buyer_address']).strip() if pd.notna(row['buyer_address']) else ""
        buyer_registration_type = str(row['buyer_registration_type']).strip() if pd.notna(row['buyer_registration_type']) else "Registered"

        if buyer_registration_type == "Registered" and not buyer_ntn_cnic:
            validation_errors.append(
                f"Row {excel_row} (Invoice {invoice_number}): "
                f"buyer NTN/CNIC is required for registered buyers."
            )
            continue

        if not buyer_business_name:
            validation_errors.append(
                f"Row {excel_row} (Invoice {invoice_number}): "
                f"buyer business name is required."
            )
            continue

        if not buyer_province:
            validation_errors.append(
                f"Row {excel_row} (Invoice {invoice_number}): "
                f"buyer province is required."
            )
            continue

        if not buyer_address:
            validation_errors.append(
                f"Row {excel_row} (Invoice {invoice_number}): "
                f"buyer address is required."
            )
            continue

        if quantity <= 0:
            validation_errors.append(
                f"Row {excel_row} (Invoice {invoice_number}): "
                f"quantity must be greater than 0."
            )
            continue

        if value_sales_excluding_st <= 0:
            validation_errors.append(
                f"Row {excel_row} (Invoice {invoice_number}): "
                f"value_sales_excluding_st must be greater than 0."
            )
            continue

        # Fixed/Retail Price >= Value Excl validation enforced only for 3rd Schedule Goods
        if saved_item.transaction_type == '3rd Schedule Goods' and fixed_notified_value_or_retail_price < value_sales_excluding_st:
            validation_errors.append(
                f"Row {excel_row} (Invoice {invoice_number}): "
                f"fixed_notified_value_or_retail_price ({fixed_notified_value_or_retail_price}) "
                f"must be equal to or greater than value_sales_excluding_st ({value_sales_excluding_st})."
            )
            continue

        if discount > value_sales_excluding_st:
            validation_errors.append(
                f"Row {excel_row} (Invoice {invoice_number}): "
                f"discount ({discount}) cannot exceed value_sales_excluding_st ({value_sales_excluding_st})."
            )
            continue

        if invoice_number not in invoice_groups:
            invoice_groups[invoice_number] = {
                "external_id": invoice_number,
                "invoice_type": str(row['invoice_type']).strip() if pd.notna(row['invoice_type']) else "Sale Invoice",
                "invoice_date": invoice_date_str,
                "transaction_type_id": saved_item.transaction_type or "01",
                "seller_ntn_cnic": seller_info.get("seller_ntn_cnic", ""),
                "seller_business_name": seller_info.get("seller_business_name", ""),
                "seller_province": seller_info.get("seller_province", ""),
                "seller_address": seller_info.get("seller_address", ""),
                "buyer_ntn_cnic": buyer_ntn_cnic,
                "buyer_business_name": buyer_business_name,
                "buyer_province": buyer_province,
                "buyer_address": buyer_address,
                "buyer_registration_type": buyer_registration_type,
                "invoice_ref_no": "",
                "scenario_id": "",
                "items": [],
                "environment": Environment.PRODUCTION,
                "income_tax": income_tax,
            }
        else:
            existing_date = invoice_groups[invoice_number]["invoice_date"]
            if invoice_date_str != existing_date:
                validation_errors.append(
                    f"Row {excel_row} (Invoice {invoice_number}): "
                    f"invoice_date '{invoice_date_str}' does not match row's date '{existing_date}'."
                )
                continue

        invoice_groups[invoice_number]["items"].append(item)

    if validation_errors:
        error_summary = f"Found {len(validation_errors)} validation error(s):\n" + "\n".join(validation_errors[:5])
        if len(validation_errors) > 5:
            error_summary += f"\n... and {len(validation_errors) - 5} more errors"
        raise ValueError(error_summary)

    return list(invoice_groups.values())
