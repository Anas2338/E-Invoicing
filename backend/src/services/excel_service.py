"""
Excel service for template generation, parsing, and validation.
"""
from typing import Optional, Tuple
from datetime import datetime, date, time
import pandas as pd
from io import BytesIO
from pathlib import Path
from uuid import UUID
from sqlmodel import Session, select
import logging

from src.models.excel_upload_session import ExcelUploadSession, ExcelUploadProcessingStatus
from src.utils.excel_validator import ExcelValidator

logger = logging.getLogger(__name__)


class ExcelService:
    """Service for Excel file operations."""

    TEMPLATE_COLUMNS = [
        # Invoice identification
        "invoice_number",
        "invoice_type",
        "invoice_date",

        # Seller information
        "seller_ntn_cnic",
        "seller_business_name",
        "seller_province",
        "seller_address",

        # Buyer information
        "buyer_ntn_cnic",
        "buyer_business_name",
        "buyer_province",
        "buyer_address",
        "buyer_registration_type",

        # Item details - matching manual sale form exactly
        "hs_code",
        "product_description",
        "tax_rate",
        "uom",
        "quantity",
        "total_values",
        "value_sales_excluding_st",
        "fixed_notified_value_or_retail_price",
        "sales_tax_applicable",
        "sales_tax_withheld_at_source",
        "extra_tax",
        "further_tax",
        "sro_schedule_no",
        "fed_payable",
        "discount",
        "sale_type",
        "sro_item_serial_no",

        # Optional fields
        "invoice_ref_no",
        "scenario_id",

        # Scheduling
        "scheduled_date",
        "scheduled_time",

        # Environment
        "environment",

        # Status fields (auto-filled by system)
        "status",
        "reason"
    ]

    def __init__(self, db: Session):
        """
        Initialize Excel service.

        Args:
            db: Database session
        """
        self.db = db
        self.validator = ExcelValidator()

    def generate_excel_template(self) -> BytesIO:
        """
        Generate Excel template with predefined headers.

        Returns:
            BytesIO object containing Excel file
        """
        # Create DataFrame with column headers
        df = pd.DataFrame(columns=self.TEMPLATE_COLUMNS)

        # Add sample row with instructions
        sample_row = {
            # Invoice identification
            "invoice_number": "INV-001",
            "invoice_type": "Sale Invoice",
            "invoice_date": "2026-04-10",

            # Seller information
            "seller_ntn_cnic": "1234567",
            "seller_business_name": "ABC Company",
            "seller_province": "PUNJAB",
            "seller_address": "123 Main Street, Lahore",

            # Buyer information
            "buyer_ntn_cnic": "7654321",
            "buyer_business_name": "XYZ Corporation",
            "buyer_province": "SINDH",
            "buyer_address": "456 Business Ave, Karachi",
            "buyer_registration_type": "Registered",

            # Item details - matching manual sale form
            "hs_code": "8471.30.00",
            "product_description": "Laptop Computer",
            "tax_rate": "18",
            "uom": "NOS",
            "quantity": "1",
            "total_values": "118000",
            "value_sales_excluding_st": "100000",
            "fixed_notified_value_or_retail_price": "0",
            "sales_tax_applicable": "18000",
            "sales_tax_withheld_at_source": "0",
            "extra_tax": "0",
            "further_tax": "0",
            "sro_schedule_no": "",
            "fed_payable": "0",
            "discount": "0",
            "sale_type": "01",
            "sro_item_serial_no": "",

            # Optional fields
            "invoice_ref_no": "",
            "scenario_id": "SN001",

            # Scheduling
            "scheduled_date": "2026-04-10",
            "scheduled_time": "10:00",

            # Environment
            "environment": "SANDBOX",

            # Status fields (auto-filled)
            "status": "",
            "reason": ""
        }
        df = pd.concat([df, pd.DataFrame([sample_row])], ignore_index=True)

        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Invoices')

            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Invoices']

            # Set column widths for better readability
            from openpyxl.utils import get_column_letter

            column_widths = [
                15,  # invoice_number
                15,  # invoice_type
                12,  # invoice_date
                15,  # seller_ntn_cnic
                25,  # seller_business_name
                15,  # seller_province
                30,  # seller_address
                15,  # buyer_ntn_cnic
                25,  # buyer_business_name
                15,  # buyer_province
                30,  # buyer_address
                20,  # buyer_registration_type
                15,  # hs_code
                30,  # product_description
                10,  # tax_rate
                10,  # uom
                10,  # quantity
                15,  # total_values
                20,  # value_sales_excluding_st
                25,  # fixed_notified_value_or_retail_price
                20,  # sales_tax_applicable
                25,  # sales_tax_withheld_at_source
                12,  # extra_tax
                12,  # further_tax
                18,  # sro_schedule_no
                12,  # fed_payable
                10,  # discount
                12,  # sale_type
                18,  # sro_item_serial_no
                15,  # invoice_ref_no
                12,  # scenario_id
                15,  # scheduled_date
                15,  # scheduled_time
                12,  # environment
                12,  # status
                30   # reason
            ]

            for idx, width in enumerate(column_widths, start=1):
                col_letter = get_column_letter(idx)
                worksheet.column_dimensions[col_letter].width = width

        output.seek(0)
        return output

    def validate_excel_structure(self, file_source: BytesIO | str) -> Tuple[bool, Optional[str]]:
        """
        Validate Excel file structure.

        Args:
            file_source: Path to Excel file or BytesIO object

        Returns:
            Tuple of (is_valid, error_message)
        """
        return self.validator.validate_excel_structure(file_source)

    def check_duplicate_invoices(self, file_source: BytesIO | str) -> Tuple[bool, Optional[str]]:
        """
        Check for duplicate invoice numbers in Excel file.

        Args:
            file_source: Path to Excel file or BytesIO object

        Returns:
            Tuple of (is_valid, error_message)
        """
        return self.validator.validate_no_duplicate_invoices(file_source)

    def parse_excel_file(self, file_source: BytesIO | str) -> list[dict]:
        """
        Parse Excel file and return invoice data.

        Args:
            file_source: Path to Excel file or BytesIO object

        Returns:
            List of invoice dictionaries with FBR-compliant structure

        Raises:
            ValueError: If Excel file is invalid
            MemoryError: If file is too large to process in memory
        """
        try:
            # Read Excel file (pandas handles both str and BytesIO)
            # Wrap in try-except to catch memory errors during large file parsing
            try:
                df = pd.read_excel(file_source, engine='openpyxl')
            except MemoryError:
                raise MemoryError(
                    "File is too large to process in memory. "
                    "Please reduce the file size or split into smaller batches (max 1,000 rows)."
                )

            # Remove sample row if present (first row with "INV-001")
            if not df.empty and df.iloc[0]['invoice_number'] == 'INV-001':
                df = df.iloc[1:]

            # Remove empty rows
            df = df.dropna(subset=['invoice_number'])

            # Convert to list of dictionaries
            invoices = []
            validation_errors = []  # Collect all validation errors
            for row_idx, row in df.iterrows():
                # Parse numeric fields
                quantity = float(row['quantity']) if pd.notna(row['quantity']) else 0
                total_values = float(row['total_values']) if pd.notna(row['total_values']) else 0
                value_sales_excluding_st = float(row['value_sales_excluding_st']) if pd.notna(row['value_sales_excluding_st']) else 0
                fixed_notified_value_or_retail_price = float(row['fixed_notified_value_or_retail_price']) if pd.notna(row['fixed_notified_value_or_retail_price']) else 0
                sales_tax_applicable = float(row['sales_tax_applicable']) if pd.notna(row['sales_tax_applicable']) else 0
                sales_tax_withheld_at_source = float(row['sales_tax_withheld_at_source']) if pd.notna(row['sales_tax_withheld_at_source']) else 0
                extra_tax = float(row['extra_tax']) if pd.notna(row['extra_tax']) else 0
                further_tax = float(row['further_tax']) if pd.notna(row['further_tax']) else 0
                fed_payable = float(row['fed_payable']) if pd.notna(row['fed_payable']) else 0
                discount = float(row['discount']) if pd.notna(row['discount']) else 0

                # Parse invoice_date to YYYY-MM-DD format
                invoice_date_str = ""
                if pd.notna(row['invoice_date']):
                    try:
                        invoice_date_parsed = pd.to_datetime(row['invoice_date'])
                        invoice_date_str = invoice_date_parsed.strftime('%Y-%m-%d')
                    except:
                        invoice_date_str = str(row['invoice_date']).strip()

                # Build FBR-compliant invoice data structure
                invoice_data = {
                    "invoice_number": str(row['invoice_number']).strip(),
                    "invoice_type": str(row['invoice_type']).strip() if pd.notna(row['invoice_type']) else "Sale Invoice",
                    "invoice_date": invoice_date_str,

                    # Seller information
                    "seller_ntn_cnic": str(row['seller_ntn_cnic']).strip() if pd.notna(row['seller_ntn_cnic']) else "",
                    "seller_business_name": str(row['seller_business_name']).strip() if pd.notna(row['seller_business_name']) else "",
                    "seller_province": str(row['seller_province']).strip() if pd.notna(row['seller_province']) else "",
                    "seller_address": str(row['seller_address']).strip() if pd.notna(row['seller_address']) else "",

                    # Buyer information
                    "buyer_ntn_cnic": str(row['buyer_ntn_cnic']).strip() if pd.notna(row['buyer_ntn_cnic']) else "",
                    "buyer_business_name": str(row['buyer_business_name']).strip() if pd.notna(row['buyer_business_name']) else "",
                    "buyer_province": str(row['buyer_province']).strip() if pd.notna(row['buyer_province']) else "",
                    "buyer_address": str(row['buyer_address']).strip() if pd.notna(row['buyer_address']) else "",
                    "buyer_registration_type": str(row['buyer_registration_type']).strip() if pd.notna(row['buyer_registration_type']) else "Registered",

                    # Item details (single item per row) - matching manual sale form structure
                    "items": [{
                        "hs_code": str(row['hs_code']).strip() if pd.notna(row['hs_code']) else "",
                        "product_description": str(row['product_description']).strip() if pd.notna(row['product_description']) else "",
                        "rate": str(row['tax_rate']).strip() if pd.notna(row['tax_rate']) else "18",
                        "uom": str(row['uom']).strip() if pd.notna(row['uom']) else "NOS",
                        "quantity": quantity,
                        "total_values": total_values,
                        "value_sales_excluding_st": value_sales_excluding_st,
                        "fixed_notified_value_or_retail_price": fixed_notified_value_or_retail_price,
                        "sales_tax_applicable": sales_tax_applicable,
                        "sales_tax_withheld_at_source": sales_tax_withheld_at_source,
                        "extra_tax": extra_tax,
                        "further_tax": further_tax,
                        "sro_schedule_no": str(row['sro_schedule_no']).strip() if pd.notna(row['sro_schedule_no']) else "",
                        "fed_payable": fed_payable,
                        "discount": discount,
                        "sale_type": str(row['sale_type']).strip() if pd.notna(row['sale_type']) else "01",
                        "sro_item_serial_no": str(row['sro_item_serial_no']).strip() if pd.notna(row['sro_item_serial_no']) else "",
                    }],

                    # Optional fields
                    "invoice_ref_no": str(row['invoice_ref_no']).strip() if pd.notna(row['invoice_ref_no']) else "",
                    "scenario_id": str(row['scenario_id']).strip() if pd.notna(row['scenario_id']) else "",

                    # Environment
                    "environment": str(row['environment']).strip() if pd.notna(row['environment']) else "SANDBOX",
                }

                # SECURITY: Validate invoice data before adding to list
                from src.utils.invoice_validator import InvoiceValidator
                is_valid, validation_error = InvoiceValidator.validate_invoice_data(invoice_data)

                if not is_valid:
                    logger.warning(f"Invoice validation failed for {invoice_data['invoice_number']}: {validation_error}")
                    # Collect validation error with row number (Excel row = pandas index + 2, accounting for header)
                    validation_errors.append(f"Row {row_idx + 2} (Invoice {invoice_data['invoice_number']}): {validation_error}")
                    continue

                # Parse scheduling information
                scheduled_date = pd.to_datetime(row['scheduled_date']).date() if pd.notna(row['scheduled_date']) else None
                scheduled_time_str = str(row['scheduled_time']).strip() if pd.notna(row['scheduled_time']) else "00:00"

                # Handle time parsing (could be "10:00" or datetime)
                try:
                    if isinstance(row['scheduled_time'], time):
                        scheduled_time = row['scheduled_time']
                    else:
                        scheduled_time = pd.to_datetime(scheduled_time_str, format='%H:%M').time()
                except:
                    scheduled_time = time(0, 0)

                invoices.append({
                    "invoice_data": invoice_data,
                    "scheduled_date": scheduled_date,
                    "scheduled_time": scheduled_time,
                })

            # If we have validation errors, report them to the user
            if validation_errors:
                error_summary = f"Found {len(validation_errors)} validation error(s):\n" + "\n".join(validation_errors[:5])
                if len(validation_errors) > 5:
                    error_summary += f"\n... and {len(validation_errors) - 5} more errors"
                raise ValueError(error_summary)

            return invoices

        except Exception as e:
            raise ValueError(f"Error parsing Excel file: {str(e)}")

    def check_concurrent_upload(self, user_id: UUID) -> Optional[ExcelUploadSession]:
        """
        Check if user has a concurrent upload in progress.

        Args:
            user_id: User UUID

        Returns:
            ExcelUploadSession if found, None otherwise
        """
        statement = select(ExcelUploadSession).where(
            ExcelUploadSession.user_id == user_id,
            ExcelUploadSession.processing_status == ExcelUploadProcessingStatus.PROCESSING
        )
        return self.db.exec(statement).first()

    def update_excel_with_status(self, file_path: str, invoice_updates: list[dict]) -> None:
        """
        Update Excel file with invoice status and reason.

        Args:
            file_path: Path to Excel file
            invoice_updates: List of dicts with invoice_number, status, reason

        Raises:
            ValueError: If Excel file cannot be updated
        """
        try:
            # Read existing Excel file
            abs_path = Path.cwd() / file_path
            df = pd.read_excel(abs_path, engine='openpyxl')

            # Create lookup dictionary for updates
            updates_dict = {
                update['invoice_number']: {
                    'status': update['status'],
                    'reason': update.get('reason', '')
                }
                for update in invoice_updates
            }

            # Update status and reason columns
            for idx, row in df.iterrows():
                invoice_num = str(row['invoice_number']).strip()
                if invoice_num in updates_dict:
                    df.at[idx, 'status'] = updates_dict[invoice_num]['status']
                    df.at[idx, 'reason'] = updates_dict[invoice_num]['reason']

            # Write back to Excel file
            with pd.ExcelWriter(abs_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Invoices')

                # Preserve column widths
                worksheet = writer.sheets['Invoices']
                column_widths = {
                    'A': 15, 'B': 25, 'C': 30, 'D': 12, 'E': 10,
                    'F': 15, 'G': 15, 'H': 15, 'I': 30,
                }
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width

        except Exception as e:
            raise ValueError(f"Error updating Excel file: {str(e)}")

    def generate_excel_from_database(self, invoices: list) -> BytesIO:
        """
        Generate Excel file from database invoice records.
        Used for exporting processed invoices with their status and results.

        Args:
            invoices: List of AutomationInvoice objects

        Returns:
            BytesIO object containing Excel file
        """
        # Prepare data rows
        rows = []
        for invoice in invoices:
            invoice_data = invoice.invoice_data

            # Extract first item (single item per invoice in current implementation)
            item = invoice_data.get('items', [{}])[0] if invoice_data.get('items') else {}

            row = {
                # Invoice identification
                "invoice_number": invoice_data.get('invoice_number', ''),
                "invoice_type": invoice_data.get('invoice_type', ''),
                "invoice_date": invoice_data.get('invoice_date', ''),

                # Seller information
                "seller_ntn_cnic": invoice_data.get('seller_ntn_cnic', ''),
                "seller_business_name": invoice_data.get('seller_business_name', ''),
                "seller_province": invoice_data.get('seller_province', ''),
                "seller_address": invoice_data.get('seller_address', ''),

                # Buyer information
                "buyer_ntn_cnic": invoice_data.get('buyer_ntn_cnic', ''),
                "buyer_business_name": invoice_data.get('buyer_business_name', ''),
                "buyer_province": invoice_data.get('buyer_province', ''),
                "buyer_address": invoice_data.get('buyer_address', ''),
                "buyer_registration_type": invoice_data.get('buyer_registration_type', ''),

                # Item details
                "hs_code": item.get('hs_code', ''),
                "product_description": item.get('product_description', ''),
                "tax_rate": item.get('rate', ''),
                "uom": item.get('uom', ''),
                "quantity": item.get('quantity', 0),
                "total_values": item.get('total_values', 0),
                "value_sales_excluding_st": item.get('value_sales_excluding_st', 0),
                "fixed_notified_value_or_retail_price": item.get('fixed_notified_value_or_retail_price', 0),
                "sales_tax_applicable": item.get('sales_tax_applicable', 0),
                "sales_tax_withheld_at_source": item.get('sales_tax_withheld_at_source', 0),
                "extra_tax": item.get('extra_tax', 0),
                "further_tax": item.get('further_tax', 0),
                "sro_schedule_no": item.get('sro_schedule_no', ''),
                "fed_payable": item.get('fed_payable', 0),
                "discount": item.get('discount', 0),
                "sale_type": item.get('sale_type', ''),
                "sro_item_serial_no": item.get('sro_item_serial_no', ''),

                # Optional fields
                "invoice_ref_no": invoice_data.get('invoice_ref_no', ''),
                "scenario_id": invoice_data.get('scenario_id', ''),

                # Scheduling
                "scheduled_date": invoice.scheduled_date.isoformat() if invoice.scheduled_date else '',
                "scheduled_time": invoice.scheduled_time.strftime('%H:%M') if invoice.scheduled_time else '',

                # Environment
                "environment": invoice_data.get('environment', ''),

                # Status fields (from processing results)
                "status": invoice.status.value if invoice.status else '',
                "reason": self._format_reason(invoice)
            }
            rows.append(row)

        # Create DataFrame
        df = pd.DataFrame(rows, columns=self.TEMPLATE_COLUMNS)

        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Invoices')

            # Get workbook and worksheet
            worksheet = writer.sheets['Invoices']

            # Set column widths (same as template)
            from openpyxl.utils import get_column_letter

            column_widths = [
                15, 15, 12, 15, 25, 15, 30, 15, 25, 15, 30, 20,
                15, 30, 10, 10, 10, 15, 20, 25, 20, 25, 12, 12,
                18, 12, 10, 12, 18, 15, 12, 15, 15, 12, 12, 30
            ]

            for idx, width in enumerate(column_widths, start=1):
                col_letter = get_column_letter(idx)
                worksheet.column_dimensions[col_letter].width = width

        output.seek(0)
        return output

    def _format_reason(self, invoice) -> str:
        """
        Format reason text based on invoice status and results.

        Args:
            invoice: AutomationInvoice object

        Returns:
            Formatted reason string
        """
        if invoice.status.value == 'submitted':
            reason = "Successfully submitted to FBR"
            if invoice.fbr_response and invoice.fbr_response.get('reference_number'):
                reason += f" (Ref: {invoice.fbr_response['reference_number']})"
            return reason
        elif invoice.status.value == 'failed':
            return invoice.validation_errors or "Processing failed"
        elif invoice.status.value == 'expired':
            return "Scheduled time is in the past"
        elif invoice.status.value == 'pending':
            return "Waiting for scheduled time"
        elif invoice.status.value == 'validated':
            return "Validated, awaiting submission"
        else:
            return ""
